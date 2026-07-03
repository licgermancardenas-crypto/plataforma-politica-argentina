"""
Dataset Político PBA 2023-2027
Generador de Excel con 6 hojas de datos políticos de la Provincia de Buenos Aires.
Fuentes: JEPBA, HCD-BA, Senado-BA, HCDN, INDEC Censo 2022, datos.gba.gob.ar
"""

import csv, json, os, unicodedata
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))

# ══════════════════════════════════════════════════════════════════════
# 1. CARGAR INTENDENTES Y SECRETARIOS DEL SCRIPT BASE
# ══════════════════════════════════════════════════════════════════════
with open(os.path.join(BASE, 'crear_excel_intendentes.py'), encoding='utf-8') as _f:
    _content = _f.read()
_cut = _content.index('\n# ─────────────────────────────────────────────────────────────\n# ESTILOS')
exec(_content[:_cut])
# INTENDENTES (lista de 135 tuplas) y SECRETARIOS (dict) ahora disponibles

# ══════════════════════════════════════════════════════════════════════
# 2. CARGA RESULTADOS ELECTORALES 2023 (datos abiertos PBA)
# ══════════════════════════════════════════════════════════════════════
def _norm(s):
    s = s.lower().strip()
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn').replace('.','').replace('  ',' ')

# Mapeo manual de 4 nombres con discrepancia CSV ↔ INTENDENTES
_CSV_MAP = {
    '25 de mayo':                    'Veinticinco de Mayo',
    'coronel de marina l rosales':   'Coronel Rosales',
    'general lamadrid':              'General La Madrid',
    'partido de la costa':           'La Costa',
}
_norm_to_muni = {_norm(m): m for m, _, _ in INTENDENTES}
_norm_to_muni.update(_CSV_MAP)

RESULTADO_2023 = {}
_csv_path = os.path.join(BASE, 'elecciones_pba_2005_2023.csv')
if os.path.exists(_csv_path):
    with open(_csv_path, encoding='latin-1') as f:
        _rows = list(csv.DictReader(f))
    _int = [r for r in _rows if r['eleccion']=='2023' and r['cargo']=='Intendente'
            and r['lista'] not in ('VOTO EN BLANCO','VOTOS NULOS') and int(r['votos'])>0]
    _by = defaultdict(list)
    for r in _int:
        _by[_norm(r['distrito'])].append(r)
    for norm_key, listas in _by.items():
        muni = _norm_to_muni.get(norm_key)
        if not muni:
            continue
        listas_s = sorted(listas, key=lambda x: int(x['votos']), reverse=True)
        total = sum(int(r['votos']) for r in listas_s)
        w, s2 = listas_s[0], (listas_s[1] if len(listas_s)>1 else None)
        RESULTADO_2023[muni] = {
            'ganador':     w['lista'],
            'ganador_pct': round(int(w['votos'])/total*100, 1),
            'segundo':     s2['lista'] if s2 else '',
            'segundo_pct': round(int(s2['votos'])/total*100, 1) if s2 else 0,
            'margen':      round((int(w['votos'])-(int(s2['votos']) if s2 else 0))/total*100, 1),
            'votantes':    int(w['votantes_habilitados']),
        }

# ══════════════════════════════════════════════════════════════════════
# 3. CARGA POBLACIÓN CENSO 2022 (INDEC)
# ══════════════════════════════════════════════════════════════════════
_CENSO_MAP = {
    # INTENDENTES name → exact Census 2022 name (where they differ)
    'Coronel Rosales':    'Coronel de Marina Leonardo Rosales',
    'General Madariaga':  'General Juan Madariaga',
    'General Pueyrredon': 'General Pueyrredón',
    'Nueve de Julio':     '9 de Julio',
    'Veinticinco de Mayo': '25 de Mayo',
}
POBLACION_2022 = {}
_censo_path = os.path.join(BASE, 'censo2022_pba.xlsx')
if os.path.exists(_censo_path):
    _wb_c = openpyxl.load_workbook(_censo_path, read_only=True)
    _ws_c = _wb_c['Cuadro2.2']
    _pop_raw = {}
    for row in _ws_c.iter_rows(values_only=True):
        if row[0] and isinstance(row[0], int) and row[0]>1000 and row[1] and isinstance(row[3],(int,float)):
            _pop_raw[str(row[1]).strip()] = int(row[3])
    for muni, _, _ in INTENDENTES:
        census_key = _CENSO_MAP.get(muni, muni)
        POBLACION_2022[muni] = _pop_raw.get(census_key, 0)

# ══════════════════════════════════════════════════════════════════════
# 4. CONCEJALES 2023 (JEPBA – Escrutinio Definitivo)
# ══════════════════════════════════════════════════════════════════════
CONCEJALES_2023 = {}  # {municipio: [{'nombre': ..., 'partido': ...}, ...]}
_conc_path = os.path.join(BASE, 'concejales_2023.json')
if os.path.exists(_conc_path):
    with open(_conc_path, encoding='utf-8') as f:
        _raw_conc = json.load(f)
    # Build municipio → list mapping; normalize UPPERCASE district names → INTENDENTES names
    _upper_to_muni = {}
    for m, _, _ in INTENDENTES:
        _upper_to_muni[_norm(m)] = m
    for _dist_id, _data in _raw_conc.items():
        _upper = _norm(_data['distrito'])
        _muni = _upper_to_muni.get(_upper) or _upper_to_muni.get(_norm(_CSV_MAP.get(_upper, '')))
        if _muni:
            CONCEJALES_2023[_muni] = {
                'n_elige': _data.get('n_elige'),
                'concejales': _data.get('concejales', []),
            }

# ══════════════════════════════════════════════════════════════════════
# 5. DATOS ESTÁTICOS
# ══════════════════════════════════════════════════════════════════════

# ── Secciones electorales (1–8) ────────────────────────────────────
SECCION_ELECTORAL = {
    # ── 1ª Sección (norte Conurbano / GBA noroeste) ──────────────
    "Campana": 1, "Escobar": 1, "General Las Heras": 1, "General Rodríguez": 1,
    "General San Martín": 1, "Hurlingham": 1, "Ituzaingó": 1, "José C. Paz": 1,
    "Luján": 1, "Malvinas Argentinas": 1, "Marcos Paz": 1, "Mercedes": 1,
    "Merlo": 1, "Moreno": 1, "Morón": 1, "Navarro": 1, "Pilar": 1,
    "San Fernando": 1, "San Isidro": 1, "San Miguel": 1, "Suipacha": 1,
    "Tigre": 1, "Tres de Febrero": 1, "Vicente López": 1,
    # ── 2ª Sección (norte PBA) ─────────────────────────────────────
    "Arrecifes": 2, "Baradero": 2, "Capitán Sarmiento": 2, "Carmen de Areco": 2,
    "Colón": 2, "Exaltación de la Cruz": 2, "Pergamino": 2, "Ramallo": 2,
    "Rojas": 2, "Salto": 2, "San Andrés de Giles": 2, "San Antonio de Areco": 2,
    "San Nicolás": 2, "San Pedro": 2, "Zárate": 2,
    # ── 3ª Sección (sur Conurbano) ─────────────────────────────────
    "Almirante Brown": 3, "Avellaneda": 3, "Berazategui": 3, "Berisso": 3,
    "Brandsen": 3, "Cañuelas": 3, "Ensenada": 3, "Esteban Echeverría": 3,
    "Ezeiza": 3, "Florencio Varela": 3, "La Matanza": 3, "Lanús": 3,
    "Lobos": 3, "Lomas de Zamora": 3, "Magdalena": 3, "Presidente Perón": 3,
    "Punta Indio": 3, "Quilmes": 3, "San Vicente": 3,
    # ── 4ª Sección (noroeste PBA) ──────────────────────────────────
    "Alberti": 4, "Bragado": 4, "Carlos Casares": 4, "Carlos Tejedor": 4,
    "Chacabuco": 4, "Chivilcoy": 4, "Florentino Ameghino": 4, "General Arenales": 4,
    "General Pinto": 4, "General Viamonte": 4, "General Villegas": 4,
    "Hipólito Yrigoyen": 4, "Junín": 4, "Leandro N. Alem": 4, "Lincoln": 4,
    "Nueve de Julio": 4, "Pehuajó": 4, "Rivadavia": 4, "Trenque Lauquen": 4,
    # ── 5ª Sección (costa atlántica + interior este) ───────────────
    "Ayacucho": 5, "Balcarce": 5, "Castelli": 5, "Chascomús": 5, "Dolores": 5,
    "General Alvarado": 5, "General Belgrano": 5, "General Guido": 5,
    "General Lavalle": 5, "General Madariaga": 5, "General Paz": 5,
    "General Pueyrredon": 5, "La Costa": 5, "Las Flores": 5, "Lezama": 5,
    "Lobería": 5, "Maipú": 5, "Mar Chiquita": 5, "Monte": 5, "Necochea": 5,
    "Pila": 5, "Pinamar": 5, "Rauch": 5, "San Cayetano": 5, "Tandil": 5,
    "Tordillo": 5, "Villa Gesell": 5,
    # ── 6ª Sección (sudoeste PBA) ──────────────────────────────────
    "Adolfo Alsina": 6, "Adolfo Gonzales Chaves": 6, "Bahía Blanca": 6,
    "Benito Juárez": 6, "Coronel Dorrego": 6, "Coronel Pringles": 6,
    "Coronel Rosales": 6, "Coronel Suárez": 6, "Daireaux": 6,
    "General La Madrid": 6, "Guaminí": 6, "Laprida": 6, "Monte Hermoso": 6,
    "Patagones": 6, "Pellegrini": 6, "Puán": 6, "Saavedra": 6,
    "Salliqueló": 6, "Tornquist": 6, "Tres Arroyos": 6, "Tres Lomas": 6,
    "Villarino": 6,
    # ── 7ª Sección (centro PBA) ────────────────────────────────────
    "Azul": 7, "Bolívar": 7, "General Alvear": 7, "Olavarría": 7,
    "Roque Pérez": 7, "Saladillo": 7, "Tapalqué": 7, "Veinticinco de Mayo": 7,
    # ── 8ª Sección (Capital – La Plata) ───────────────────────────
    "La Plata": 8,
}

SECCION_NOMBRE = {
    1: "1ª – Norte Conurbano",
    2: "2ª – Norte PBA",
    3: "3ª – Sur Conurbano",
    4: "4ª – Noroeste PBA",
    5: "5ª – Costa Atlántica",
    6: "6ª – Sudoeste PBA",
    7: "7ª – Centro PBA",
    8: "8ª – Capital (La Plata)",
}

# ── Diputados Provinciales (Honorable Cámara de Diputados PBA) ─────
# Fuente: hcdiputados-ba.gov.ar – composición julio 2026
# Sección: número 1-8 | Bloque al momento del relevamiento
DIPUTADOS_PROV = [
    # FUERZA PATRIA (oficialismo K – Unión por la Patria)
    ("Esteban Alejandro Acerbo",       6, "Fuerza Patria"),
    ("Maria Laura Aloisi",             7, "Fuerza Patria"),
    ("Soledad Antonia Alonso",         1, "Fuerza Patria"),
    ("Maite Milagros Alvado",          6, "Fuerza Patria"),
    ("Juan Ariel Archanco",            8, "Fuerza Patria"),
    ("Ana Luz Balor",                  1, "Fuerza Patria"),
    ("Romina Melisa Barreiro",         3, "Fuerza Patria"),
    ("Marcela María Amelia Basualdo",  5, "Fuerza Patria"),
    ("Mariano Cascallares",            3, "Fuerza Patria"),
    ("Juan Pablo De Jesus",            5, "Fuerza Patria"),
    ("Germán Di Cesare",               4, "Fuerza Patria"),
    ("Enrique Alejandro Dichiara",     6, "Fuerza Patria"),
    ("Héctor Rubén Eslaiman",          1, "Fuerza Patria"),
    ("Evelyn Marlen Flores Yanz",      2, "Fuerza Patria"),
    ("José Renaldo Galván",            3, "Fuerza Patria"),
    ("Alexis Guerrera",                4, "Fuerza Patria"),
    ("Viviana Raquel Guzzo",           4, "Fuerza Patria"),
    ("Lucía Iañez",                    8, "Fuerza Patria"),
    ("María Eva Limone",               3, "Fuerza Patria"),
    ("Ricardo Lissalde",               7, "Fuerza Patria"),
    ("Juan Martín Malpeli",            8, "Fuerza Patria"),
    ("Mayra Soledad Mendoza",          3, "Fuerza Patria"),
    ("Leonardo José Moreno",           1, "Fuerza Patria"),
    ("Maria Silvina Nardini",          3, "Fuerza Patria"),
    ("Diego Eduardo Nanni",            2, "Fuerza Patria"),
    ("Micaela Edith Olivetto",         4, "Fuerza Patria"),
    ("Luciana Padulo",                 5, "Fuerza Patria"),
    ("Carlos Javier Puglelli",         2, "Fuerza Patria"),
    ("Ayelén Itatí Rasquetti",         3, "Fuerza Patria"),
    ("Margarita Alejandra Recalde",    1, "Fuerza Patria"),
    ("Jose Ignacio Rossi",             5, "Fuerza Patria"),
    ("María Noelia Saavedra",          1, "Fuerza Patria"),
    ("Facundo Tignanelli",             3, "Fuerza Patria"),
    ("Sofía Vannelli",                 6, "Fuerza Patria"),
    ("Roberto Ramón Vazquez",          3, "Fuerza Patria"),
    ("Luis Omar Vivona",               3, "Fuerza Patria"),
    ("Avelino Ricardo Zurro",          4, "Fuerza Patria"),
    # LA LIBERTAD AVANZA
    ("Gastón Andrés Abonjo",           5, "La Libertad Avanza"),
    ("Francisco Jorge Adorni",         8, "La Libertad Avanza"),
    ("Maximiliano Iván Bondarenko",    3, "La Libertad Avanza"),
    ("Héctor Norberto Gay",            6, "La Libertad Avanza"),
    ("Oscar Eduardo Liberman",         6, "La Libertad Avanza"),
    ("Pablo Ezequiel Morillo",         2, "La Libertad Avanza"),
    ("Constanza Moragues",             1, "La Libertad Avanza"),
    ("Agustín Romo",                   7, "La Libertad Avanza"),
    ("Nahuel Sotelo Larcher",          3, "La Libertad Avanza"),
    ("Mariela Silvina Vitale",         6, "La Libertad Avanza"),
    # PRO
    ("Natalia Miriam Blanco",          2, "PRO"),
    ("Leticia Inés Bontempo",          3, "PRO"),
    ("María Paula Bustos",             2, "PRO"),
    ("Martín Julián Endere",           7, "PRO"),
    ("Agustín Forchieri",              1, "PRO"),
    ("Maria Laura Ricchini",           4, "PRO"),
    ("Fernando Rovello",               4, "PRO"),
    ("Ana Rita Sallaberry",            1, "PRO"),
    ("María Angel Sotolano",           3, "PRO"),
    ("Julieta Elisabet Quintero Chasman", 8, "PRO"),
    ("Alejandro Daniel Rabinovich",    2, "PRO"),
    # UNIÓN Y LIBERTAD
    ("Blanca Elida Alessi",            4, "Unión y Libertad"),
    ("María Laura Fernández",          7, "Unión y Libertad"),
    ("María Salomé Jalil Toledo",      1, "Unión y Libertad"),
    ("Martín Adolfo Rozas",            4, "Unión y Libertad"),
    ("Sabrina Inés Sabat",             5, "Unión y Libertad"),
    # UCR + CAMBIO FEDERAL
    ("Matías Raúl Civale",             5, "UCR + Cambio Federal"),
    ("Diego Raúl Garciarena",          5, "UCR + Cambio Federal"),
    ("María Silvina Vaccarezza",       4, "UCR + Cambio Federal"),
    # UCR
    ("Valentín Miranda",               4, "UCR"),
    ("Priscila Minnaard",              6, "UCR"),
    ("Maria Alejandra Lorden",         7, "UCR"),
    # NUEVOS AIRES
    ("Gustavo Sergio Cuervo",          4, "Nuevos Aires"),
    ("Carlos Fabián Luayza Troncozo",  1, "Nuevos Aires"),
    ("Viviana Yolanda Romano",         1, "Nuevos Aires"),
    # COALICIÓN CÍVICA
    ("Romina Natalí Braga",            5, "Coalición Cívica"),
    ("Luciano David Bugallo Di Prinzio", 4, "Coalición Cívica"),
    ("José Andrés De Leo",             6, "Coalición Cívica"),
    # ESPACIO ABIERTO – HECHOS
    ("Ignacio Fernando Mateucci",      2, "Espacio Abierto – Hechos"),
    ("Manuel Passaglia",               2, "Espacio Abierto – Hechos"),
    # DERECHA POPULAR
    ("Juan José Esper Zamar",          1, "Derecha Popular"),
    # IZQUIERDA (FIT-U)
    ("Monica Leticia Schlotthauer",    3, "Izquierda Socialista – FIT-U"),
    ("Christian Castillo",             3, "PTS – FIT-U"),
]

# ── Senadores Provinciales (Cámara de Senadores PBA) ───────────────
# Fuente: senado-ba.gov.ar – composición julio 2026
SENADORES_PROV = [
    # FUERZA PATRIA
    ("Valeria Arata",                  4, "Fuerza Patria"),
    ("Sabrina Bastida",                8, "Fuerza Patria"),
    ("Sergio Alejandro Berni",         2, "Fuerza Patria"),
    ("Pedro Francisco Borgini",        8, "Fuerza Patria"),
    ("Laura Magdalena Clark",          2, "Fuerza Patria"),
    ("Fernando Gabriel Coronel",       1, "Fuerza Patria"),
    ("Amira Curi",                     3, "Fuerza Patria"),
    ("Evelyn Diaz",                    7, "Fuerza Patria"),
    ("Ayelén Durán",                   6, "Fuerza Patria"),
    ("Federico Fagioli",               3, "Fuerza Patria"),
    ("Marcelo Enrique Feliú",          6, "Fuerza Patria"),
    ("Malena Galmarini",               1, "Fuerza Patria"),
    ("Emmanuel Severo González Santalla", 3, "Fuerza Patria"),
    ("Mario Alberto Ishii",            1, "Fuerza Patria"),
    ("German Lago",                    4, "Fuerza Patria"),
    ("María Inés Laurini",             7, "Fuerza Patria"),
    ("Roxana Lopez",                   1, "Fuerza Patria"),
    ("Monica Fernanda Macha",          1, "Fuerza Patria"),
    ("Maria Rosa Martinez",            3, "Fuerza Patria"),
    ("Jorge Alberto Paredi",           5, "Fuerza Patria"),
    ("Marcos Emilio Pisano",           7, "Fuerza Patria"),
    ("Maria Fernanda Raverta",         5, "Fuerza Patria"),
    ("Adrián Carlos Santarelli",       3, "Fuerza Patria"),
    ("Diego Alberto Videla",           4, "Fuerza Patria"),
    # LA LIBERTAD AVANZA
    ("Maria Florencia Arietto",        3, "La Libertad Avanza"),
    ("Analia Balaudo",                 4, "La Libertad Avanza"),
    ("Maria Luz Bambaci",              1, "La Libertad Avanza"),
    ("Gonzalo Cabezas",                4, "La Libertad Avanza"),
    ("Carlos Nicolas Curestis",        3, "La Libertad Avanza"),
    ("Matias Miguel De Urraza",        5, "La Libertad Avanza"),
    ("Maria Cecilia Martinez",         5, "La Libertad Avanza"),
    ("Luciano Emanuel Olivera",        1, "La Libertad Avanza"),
    ("Betina Clara Riva",              3, "La Libertad Avanza"),
    ("Diego Valenzuela",               1, "La Libertad Avanza"),
    # PRO
    ("Alex Campbell",                  6, "PRO"),
    ("Guillermo Tristán Montenegro",   5, "PRO"),
    ("Pablo Alexis Petrecca",          4, "PRO"),
    ("Juan Manuel Rico Zini",          2, "PRO"),
    ("Jorge Schiavone",                3, "PRO"),
    # HECHOS – UCR IDENTIDAD
    ("Marcelo Leguizamón Brown",       8, "Hechos – UCR Identidad"),
    ("Natalia Quintana",               4, "Hechos – UCR Identidad"),
    ("Maria Emilia Subiza",            2, "Hechos – UCR Identidad"),
    # UCR
    ("Nerina Daniela Neumann",         6, "UCR"),
    # UNIÓN Y LIBERTAD
    ("Carlos Francisco Kikuchi",       2, "Unión y Libertad"),
    ("Sergio Raul Vargas",             6, "Unión y Libertad"),
    ("Silvana Paola Ventura",          6, "Unión y Libertad"),
]

# ── Diputados Nacionales (Cámara de Diputados de la Nación) ────────
# Buenos Aires provincia, mandato 2023-2027 (elegidos octubre 2023)
DIPUTADOS_NAC = [
    ("Constanza María Alonso",         "Unión por la Patria"),
    ("Pablo Ansaloni",                 "La Libertad Avanza"),
    ("Daniel Fernando Arroyo",         "Unión por la Patria"),
    ("Karina Verónica Banfi",          "UCR"),
    ("Alberto Benegas Lynch",          "La Libertad Avanza"),
    ("Gabriela Besana",                "PRO"),
    ("Victoria Borrego",               "Coalición Cívica"),
    ("Santiago Cafiero",               "Unión por la Patria"),
    ("María Marcela Campagnoli",       "Coalición Cívica"),
    ("Carlos Daniel Castagneto",       "Unión por la Patria"),
    ("Christian Castillo",             "FIT-U"),
    ("María Florencia De Sensi",       "PRO"),
    ("Juan Carlos Giordano",           "FIT-U"),
    ("Vilma Ripoll",                   "FIT-U"),
    ("José Luis Espert",               "La Libertad Avanza"),
    ("Eduardo Falcone",                "MID"),
    ("Alejandro Finocchiaro",          "PRO"),
    ("Mónica Edith Frade",             "Coalición Cívica"),
    ("Daniel Gollán",                  "Unión por la Patria"),
    ("Carlos Ramiro Gutiérrez",        "Unión por la Patria"),
    ("Rogelio Iparraguirre",           "Unión por la Patria"),
    ("Pablo Juliano",                  "Democracia para Siempre"),
    ("Máximo Carlos Kirchner",         "Unión por la Patria"),
    ("Lilia Adela Lemoine",            "La Libertad Avanza"),
    ("Mónica Litza",                   "Unión por la Patria"),
    ("Juan Manuel López",              "Coalición Cívica"),
    ("Silvia Gabriela Lospennato",     "PRO"),
    ("Mónica Fernanda Macha",          "Unión por la Patria"),
    ("María Lorena Macyszyn",          "La Libertad Avanza"),
    ("Facundo Manes",                  "Democracia para Siempre"),
    ("Mario Roberto Manrique",         "Unión por la Patria"),
    ("Juan Marino",                    "Unión por la Patria"),
    ("Nicolás Massot",                 "Encuentro Federal"),
    ("Gerardo Milman",                 "PRO"),
    ("Matías Molle",                   "Unión por la Patria"),
    ("Guillermo Montenegro",           "La Libertad Avanza"),
    ("Emilio Monzó",                   "Encuentro Federal"),
    ("Roxana Monzón",                  "Unión por la Patria"),
    ("Micaela Morán",                  "Unión por la Patria"),
    ("Cecilia Moreau",                 "Unión por la Patria"),
    ("Leopoldo Raúl Guido Moreau",     "Unión por la Patria"),
    ("Marcela Marina Pagano",          "Coherencia"),
    ("Sergio Omar Palazzo",            "Unión por la Patria"),
    ("Marcela Fabiana Passo",          "Unión por la Patria"),
    ("Julio César Pereyra",            "Unión por la Patria"),
    ("Miguel Ángel Pichetto",          "Encuentro Federal"),
    ("Carolina Píparo",                "La Libertad Avanza"),
    ("Luciana Potenza",                "Unión por la Patria"),
    ("Agustina Lucrecia Propato",      "Unión por la Patria"),
    ("Fabio José Quetglas",            "UCR"),
    ("Aníbal Florencio Randazzo",      "Encuentro Federal"),
    ("Cristian Adrián Ritondo",        "PRO"),
    ("Javier Sánchez Wrba",            "PRO"),
    ("Juliana Santillán Juárez Ibrahim","La Libertad Avanza"),
    ("Diego César Santilli",           "PRO"),
    ("Santiago Santurio",              "La Libertad Avanza"),
    ("Sabrina Selva",                  "Unión por la Patria"),
    ("Vanesa Raquel Siley",            "Unión por la Patria"),
    ("María Sotolano",                 "PRO"),
    ("Margarita Rosa Stolbizer",       "Encuentro Federal"),
    ("Julia Strada",                   "Unión por la Patria"),
    ("Luis Rodolfo Tailhade",          "Unión por la Patria"),
    ("Danya Tavela",                   "Democracia para Siempre"),
    ("Victoria Tolosa Paz",            "Unión por la Patria"),
    ("Brenda Vargas Matyi",            "Unión por la Patria"),
    ("Patricia Vásquez",               "La Libertad Avanza"),
    ("Luana Volnovich",                "Unión por la Patria"),
    ("Hugo Rubén Yasky",               "Unión por la Patria"),
    ("Martín Yeza",                    "PRO"),
    ("Natalia Zaracho",                "Unión por la Patria"),
]

# ── Senadores Nacionales – Buenos Aires provincia ──────────────────
# Mandato 2023-2029 (elegidos octubre 2023)
SENADORES_NAC = [
    ("Maximiliano Abad",        "UCR"),
    ("Eduardo 'Wado' de Pedro", "Unión por la Patria"),
    ("Juliana Di Tullio",       "Unión por la Patria"),
]

# ══════════════════════════════════════════════════════════════════════
# 5. HELPERS DE ESTILO
# ══════════════════════════════════════════════════════════════════════
def _side():
    return Side(style='thin', color='BDBDBD')

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def _font(bold=False, size=11, color='000000', italic=False):
    return Font(name='Calibri', bold=bold, size=size, color=color, italic=italic)

def _align(h='left', wrap=False):
    return Alignment(horizontal=h, vertical='center', wrap_text=wrap)

def header_style(cell, color='1F3864'):
    cell.fill = _fill(color)
    cell.font = _font(bold=True, size=11, color='FFFFFF')
    cell.alignment = _align('center')
    cell.border = _border()

def data_style(cell, bg=None, wrap=False, bold=False, center=False):
    if bg:
        cell.fill = _fill(bg)
    cell.font = _font(bold=bold)
    cell.alignment = _align('center' if center else 'left', wrap=wrap)
    cell.border = _border()

COLOR_UP   = 'E8F5E9'   # verde – Unión por la Patria
COLOR_JXC  = 'E3F2FD'   # azul  – JxC / PRO / LLA / oposición
COLOR_VEC  = 'FFF9C4'   # amarillo – vecinales
COLOR_LLA  = 'FFF3E0'   # naranja pálido – La Libertad Avanza
COLOR_FP   = 'E8F5E9'   # verde – Fuerza Patria (mismo que UP)

BLOQUE_COLOR = {
    'Fuerza Patria':      COLOR_FP,
    'Unión por la Patria': COLOR_UP,
    'La Libertad Avanza': COLOR_LLA,
    'PRO':                COLOR_JXC,
    'UCR':                'EDE7F6',
    'UCR + Cambio Federal': 'EDE7F6',
    'Unión y Libertad':   'F3E5F5',
    'Coalición Cívica':   'E1F5FE',
    'Encuentro Federal':  'FCE4EC',
    'Democracia para Siempre': 'F9FBE7',
    'FIT-U':              'FFEBEE',
    'PTS – FIT-U':        'FFEBEE',
    'Izquierda Socialista – FIT-U': 'FFEBEE',
    'Espacio Abierto – Hechos': 'F1F8E9',
    'Hechos – UCR Identidad': 'F1F8E9',
    'Nuevos Aires':       'FFF8E1',
    'Derecha Popular':    'EFEBE9',
    'MID':                'ECEFF1',
    'Coherencia':         'ECEFF1',
}

def get_color(partido_o_bloque):
    for k, v in BLOQUE_COLOR.items():
        if k.lower() in partido_o_bloque.lower():
            return v
    if 'Patria' in partido_o_bloque or 'peronism' in partido_o_bloque.lower():
        return COLOR_UP
    if 'libertad' in partido_o_bloque.lower() or 'avanza' in partido_o_bloque.lower():
        return COLOR_LLA
    if 'juntos' in partido_o_bloque.lower() or 'cambio' in partido_o_bloque.lower():
        return COLOR_JXC
    if 'vecinal' in partido_o_bloque.lower():
        return COLOR_VEC
    return None

def set_col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ══════════════════════════════════════════════════════════════════════
# 6. CREAR WORKBOOK
# ══════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════
# HOJA 1: INTENDENTES (con sección, población, resultados 2023)
# ══════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Intendentes"
ws1.freeze_panes = "A2"
ws1.row_dimensions[1].height = 22

COLS_H1 = [
    ("#",              4),
    ("Municipio",     22),
    ("Intendente/a",  26),
    ("Partido / Coalición", 28),
    ("Sección Electoral", 22),
    ("Población 2022", 16),
    ("Ganador 2023",  28),
    ("% Ganador",     11),
    ("2do Lugar",     28),
    ("% 2do",         10),
    ("Margen %",      10),
    ("Padrón 2023",   14),
]
for col, (title, width) in enumerate(COLS_H1, 1):
    cell = ws1.cell(row=1, column=col, value=title)
    header_style(cell)
    set_col_width(ws1, col, width)

for i, (muni, intendente, partido) in enumerate(INTENDENTES, 1):
    row = i + 1
    seccion_n = SECCION_ELECTORAL.get(muni, '')
    seccion_str = f"{seccion_n}ª" if seccion_n else ''
    pop = POBLACION_2022.get(muni, 0)
    res = RESULTADO_2023.get(muni, {})
    bg = get_color(partido)

    vals = [
        i,
        muni,
        intendente,
        partido,
        seccion_str,
        pop if pop else '',
        res.get('ganador', ''),
        res.get('ganador_pct', ''),
        res.get('segundo', ''),
        res.get('segundo_pct', ''),
        res.get('margen', ''),
        res.get('votantes', ''),
    ]
    for col, val in enumerate(vals, 1):
        cell = ws1.cell(row=row, column=col, value=val)
        is_num = col in (1, 6, 8, 10, 11, 12)
        data_style(cell, bg=(bg if row % 2 == 0 else None), center=is_num)

ws1.auto_filter.ref = f"A1:{get_column_letter(len(COLS_H1))}1"

# ══════════════════════════════════════════════════════════════════════
# HOJA 2: SECRETARIOS (sin cambios respecto al script base)
# ══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Secretarios")
ws2.freeze_panes = "A2"
ws2.row_dimensions[1].height = 22

COLS_H2 = [
    ("#", 4), ("Municipio", 22), ("Intendente/a", 26),
    ("Partido", 28), ("Secretario/a", 26), ("Cargo / Secretaría", 44),
]
for col, (title, width) in enumerate(COLS_H2, 1):
    cell = ws2.cell(row=1, column=col, value=title)
    header_style(cell)
    set_col_width(ws2, col, width)

mapa = {m: (i, p) for m, i, p in INTENDENTES}
counter = 1
row = 2
for municipio in sorted(SECRETARIOS.keys()):
    intendente, partido = mapa[municipio]
    bg = get_color(partido)
    for nombre, cargo in SECRETARIOS[municipio]:
        vals = [counter, municipio, intendente, partido, nombre, cargo]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            data_style(cell, bg=(bg if row%2==0 else None), wrap=(col==6))
        row += 1
        counter += 1

ws2.auto_filter.ref = f"A1:{get_column_letter(len(COLS_H2))}1"

# ══════════════════════════════════════════════════════════════════════
# HOJA 3: RESULTADOS ELECTORALES 2023 (detalle por municipio)
# ══════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Resultados 2023")
ws3.freeze_panes = "A2"
ws3.row_dimensions[1].height = 22

COLS_H3 = [
    ("#", 4), ("Municipio", 22), ("Sección", 14),
    ("Intendente Electo", 26), ("Partido Ganador", 30),
    ("% Votos", 10), ("2do Partido", 30), ("% 2do", 9),
    ("Margen %", 10), ("Padrón", 12), ("Resultado", 18),
]
for col, (title, width) in enumerate(COLS_H3, 1):
    cell = ws3.cell(row=1, column=col, value=title)
    header_style(cell)
    set_col_width(ws3, col, width)

def clasif_resultado(pct, margen):
    if margen >= 20: return "Victoria amplia"
    if margen >= 10: return "Victoria clara"
    if margen >= 5:  return "Ventaja moderada"
    return "Elección reñida"

for i, (muni, intendente, partido) in enumerate(INTENDENTES, 1):
    row = i + 1
    res = RESULTADO_2023.get(muni, {})
    seccion_n = SECCION_ELECTORAL.get(muni, '')
    bg = get_color(partido)
    pct = res.get('ganador_pct', 0)
    mrg = res.get('margen', 0)
    vals = [
        i,
        muni,
        f"{seccion_n}ª" if seccion_n else '',
        intendente,
        res.get('ganador', partido),
        pct,
        res.get('segundo', ''),
        res.get('segundo_pct', ''),
        mrg,
        res.get('votantes', ''),
        clasif_resultado(pct, mrg) if res else '',
    ]
    for col, val in enumerate(vals, 1):
        cell = ws3.cell(row=row, column=col, value=val)
        is_num = col in (1, 6, 8, 9, 10)
        data_style(cell, bg=(bg if row%2==0 else None), center=is_num)

ws3.auto_filter.ref = f"A1:{get_column_letter(len(COLS_H3))}1"

# ══════════════════════════════════════════════════════════════════════
# HOJA 4: LEGISLADORES PROVINCIALES
# ══════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Legisladores Provinciales")
ws4.freeze_panes = "A2"
ws4.row_dimensions[1].height = 22

COLS_H4 = [
    ("#", 4), ("Nombre", 36), ("Cámara", 16),
    ("Sección Elect.", 18), ("Bloque", 30),
]
for col, (title, width) in enumerate(COLS_H4, 1):
    cell = ws4.cell(row=1, column=col, value=title)
    header_style(cell, color='1A237E')
    set_col_width(ws4, col, width)

all_legislators = (
    [(n, "Diputados PBA", s, b) for n, s, b in DIPUTADOS_PROV] +
    [(n, "Senado PBA",    s, b) for n, s, b in SENADORES_PROV]
)
for i, (nombre, camara, seccion, bloque) in enumerate(sorted(all_legislators, key=lambda x: (x[1],x[2],x[0])), 1):
    row = i + 1
    bg = get_color(bloque)
    seccion_str = f"{seccion}ª – {SECCION_NOMBRE.get(seccion,'')}"
    vals = [i, nombre, camara, seccion_str, bloque]
    for col, val in enumerate(vals, 1):
        cell = ws4.cell(row=row, column=col, value=val)
        data_style(cell, bg=(bg if row%2==0 else None), center=(col==1))

ws4.auto_filter.ref = f"A1:{get_column_letter(len(COLS_H4))}1"

# ══════════════════════════════════════════════════════════════════════
# HOJA 5: LEGISLADORES NACIONALES (Buenos Aires provincia)
# ══════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Legisladores Nacionales")
ws5.freeze_panes = "A2"
ws5.row_dimensions[1].height = 22

COLS_H5 = [
    ("#", 4), ("Nombre", 40), ("Cámara", 22),
    ("Bloque / Partido", 32), ("Mandato", 16),
]
for col, (title, width) in enumerate(COLS_H5, 1):
    cell = ws5.cell(row=1, column=col, value=title)
    header_style(cell, color='4A148C')
    set_col_width(ws5, col, width)

all_nac = (
    [(n, "Diputados Nación", b, "2023-2027") for n, b in DIPUTADOS_NAC] +
    [(n, "Senado Nación",    b, "2023-2029") for n, b in SENADORES_NAC]
)
for i, (nombre, camara, bloque, mandato) in enumerate(sorted(all_nac, key=lambda x:(x[1],x[0])), 1):
    row = i + 1
    bg = get_color(bloque)
    vals = [i, nombre, camara, bloque, mandato]
    for col, val in enumerate(vals, 1):
        cell = ws5.cell(row=row, column=col, value=val)
        data_style(cell, bg=(bg if row%2==0 else None), center=(col in (1,5)))

ws5.auto_filter.ref = f"A1:{get_column_letter(len(COLS_H5))}1"

# ══════════════════════════════════════════════════════════════════════
# HOJA 6: CONCEJALES 2023 (electos en elecciones generales oct. 2023)
# ══════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Concejales 2023")
ws6.freeze_panes = "A2"
ws6.row_dimensions[1].height = 22

COLS_H6 = [
    ("#",                  5),
    ("Municipio",         22),
    ("Sección Electoral", 22),
    ("N° Elige",           9),
    ("N° en Municipio",   10),
    ("Nombre Concejal",   36),
    ("Partido / Lista",   42),
    ("Bloque Político",   24),
]
for col, (title, width) in enumerate(COLS_H6, 1):
    cell = ws6.cell(row=1, column=col, value=title)
    header_style(cell, color='1A237E')
    set_col_width(ws6, col, width)

def _bloque_conc(partido):
    p = partido.upper()
    if any(k in p for k in ('UNION POR LA PATRIA', 'FRENTE DE TODOS', 'PERONISMO', 'KIRCHNER',
                             'TODOS POR', 'FRENTE RENOVADOR', 'PJ ')):
        return 'Unión por la Patria'
    if any(k in p for k in ('LA LIBERTAD AVANZA', 'LIBERTAD AVANZA', 'LLA')):
        return 'La Libertad Avanza'
    if any(k in p for k in ('JUNTOS POR EL CAMBIO', 'JUNTOS SOMOS', 'PRO ',
                             'JUNTOS POR', 'PROPUESTA REPUBLICANA')):
        return 'Juntos por el Cambio'
    if any(k in p for k in ('UCR ', 'UNION CIVICA RADICAL', 'RADICALISMO')):
        return 'UCR'
    if any(k in p for k in ('FIT', 'IZQUIERDA')):
        return 'FIT-U'
    return 'Vecinal / Local'

_h6_row = 2
_h6_count = 0
for muni, _, _ in INTENDENTES:
    data = CONCEJALES_2023.get(muni)
    if not data:
        continue
    seccion_n = SECCION_ELECTORAL.get(muni, '')
    seccion_label = SECCION_NOMBRE.get(seccion_n, '') if seccion_n else ''
    n_elige = data['n_elige']
    for idx_c, conc in enumerate(data['concejales'], 1):
        bloque = _bloque_conc(conc['partido'])
        bg = get_color(bloque)
        vals = [
            _h6_count + 1,
            muni,
            seccion_label,
            n_elige,
            idx_c,
            conc['nombre'],
            conc['partido'],
            bloque,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws6.cell(row=_h6_row, column=col, value=val)
            data_style(cell,
                       bg=(bg if _h6_row % 2 == 0 else None),
                       center=(col in (1, 4, 5)))
        _h6_count += 1
        _h6_row += 1

ws6.auto_filter.ref = f"A1:{get_column_letter(len(COLS_H6))}1"

# ══════════════════════════════════════════════════════════════════════
# HOJA 7: RESUMEN / DASHBOARD
# ══════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Resumen")
ws7.column_dimensions['A'].width = 46
ws7.column_dimensions['B'].width = 22
ws7.row_dimensions[1].height = 34

def _sec(ws, row, label):
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    cell.fill = _fill('1F3864')
    cell.alignment = _align('left')
    ws.merge_cells(f'A{row}:B{row}')

def _row(ws, row, label, value, bg=None):
    ca = ws.cell(row=row, column=1, value=label)
    cb = ws.cell(row=row, column=2, value=value)
    ca.font = _font(size=11)
    cb.font = _font(bold=True, size=11)
    if bg:
        ca.fill = _fill(bg)
        cb.fill = _fill(bg)
    ca.border = _border()
    cb.border = _border()
    ca.alignment = _align('left')
    cb.alignment = _align('center')

# Title
t = ws7.cell(row=1, column=1, value="DATASET POLÍTICO PBA 2023-2027 — RESUMEN EJECUTIVO")
t.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
t.fill = _fill('0D47A1')
t.alignment = _align('left')
ws7.merge_cells('A1:B1')

r = 3
_sec(ws7, r, "COBERTURA DEL DATASET"); r+=1
_row(ws7, r, "Total municipios de Buenos Aires", 135, 'F5F5F5'); r+=1
_row(ws7, r, "Municipios con secretarios relevados", len(SECRETARIOS), 'F5F5F5'); r+=1
_row(ws7, r, "Total secretarios cargados", sum(len(v) for v in SECRETARIOS.values()), 'F5F5F5'); r+=1
_row(ws7, r, "Municipios con resultados electorales 2023", len(RESULTADO_2023), 'F5F5F5'); r+=1
_row(ws7, r, "Municipios con población Censo 2022", len(POBLACION_2022), 'F5F5F5'); r+=1
_row(ws7, r, "Municipios con concejales 2023 relevados", len(CONCEJALES_2023), 'F5F5F5'); r+=1
_row(ws7, r, "Total concejales electos 2023 (titulares)", _h6_count, 'F5F5F5'); r+=1
r+=1

_sec(ws7, r, "COMPOSICIÓN POLÍTICA – INTENDENTES 2023-2027"); r+=1
from collections import Counter
bloques = Counter()
for _, _, p in INTENDENTES:
    if 'Patria' in p: bloques['Unión por la Patria'] += 1
    elif 'Juntos' in p: bloques['Juntos por el Cambio'] += 1
    elif 'Libertad' in p or 'LLA' in p: bloques['La Libertad Avanza'] += 1
    else: bloques['Otros / Vecinales'] += 1

_row(ws7, r, "Unión por la Patria",        bloques['Unión por la Patria'],        COLOR_UP); r+=1
_row(ws7, r, "Juntos por el Cambio",       bloques['Juntos por el Cambio'],       COLOR_JXC); r+=1
_row(ws7, r, "La Libertad Avanza",         bloques['La Libertad Avanza'],         COLOR_LLA); r+=1
_row(ws7, r, "Otros / Vecinales",          bloques['Otros / Vecinales'],          COLOR_VEC); r+=1
r+=1

_sec(ws7, r, "LEGISLATURA PROVINCIAL"); r+=1
from collections import Counter as Cnt
dp_bloques = Cnt(b for _,_,b in DIPUTADOS_PROV)
sp_bloques = Cnt(b for _,_,b in SENADORES_PROV)
_row(ws7, r, "Diputados provinciales relevados", len(DIPUTADOS_PROV), 'F5F5F5'); r+=1
_row(ws7, r, "Senadores provinciales relevados", len(SENADORES_PROV), 'F5F5F5'); r+=1
_row(ws7, r, "  Fuerza Patria (Diputados)", dp_bloques.get('Fuerza Patria',0), COLOR_FP); r+=1
_row(ws7, r, "  La Libertad Avanza (Diputados)", dp_bloques.get('La Libertad Avanza',0), COLOR_LLA); r+=1
_row(ws7, r, "  PRO (Diputados)", dp_bloques.get('PRO',0), COLOR_JXC); r+=1
_row(ws7, r, "  Fuerza Patria (Senadores)", sp_bloques.get('Fuerza Patria',0), COLOR_FP); r+=1
_row(ws7, r, "  La Libertad Avanza (Senadores)", sp_bloques.get('La Libertad Avanza',0), COLOR_LLA); r+=1
_row(ws7, r, "  PRO (Senadores)", sp_bloques.get('PRO',0), COLOR_JXC); r+=1
r+=1

_sec(ws7, r, "LEGISLADORES NACIONALES – BUENOS AIRES"); r+=1
dn_bloques = Cnt(b for _,b in DIPUTADOS_NAC)
_row(ws7, r, "Diputados nacionales (mandato 2023-2027)", len(DIPUTADOS_NAC), 'F5F5F5'); r+=1
_row(ws7, r, "Senadores nacionales", len(SENADORES_NAC), 'F5F5F5'); r+=1
_row(ws7, r, "  Unión por la Patria (Diputados)", dn_bloques.get('Unión por la Patria',0), COLOR_UP); r+=1
_row(ws7, r, "  La Libertad Avanza (Diputados)", dn_bloques.get('La Libertad Avanza',0), COLOR_LLA); r+=1
_row(ws7, r, "  PRO (Diputados)", dn_bloques.get('PRO',0), COLOR_JXC); r+=1
r+=1

_sec(ws7, r, "METADATOS"); r+=1
_row(ws7, r, "Período cubierto",          "2023-2027 (asumieron dic. 2023)", 'F5F5F5'); r+=1
_row(ws7, r, "Fecha de compilación",      "Julio 2026", 'F5F5F5'); r+=1
_row(ws7, r, "Fuentes principales",
     "JEPBA · HCD-BA · Senado-BA · HCDN · INDEC Censo 2022 · datos.gba.gob.ar",
     'F5F5F5'); r+=1

# ══════════════════════════════════════════════════════════════════════
# GUARDAR
# ══════════════════════════════════════════════════════════════════════
output = os.path.join(BASE, "PBA_Dataset_Politico_2023-2027.xlsx")
wb.save(output)
print(f"✓ Excel guardado: {output}")
print(f"  Hoja 1 – Intendentes:             {len(INTENDENTES)} filas")
print(f"  Hoja 2 – Secretarios:             {sum(len(v) for v in SECRETARIOS.values())} filas ({len(SECRETARIOS)} municipios)")
print(f"  Hoja 3 – Resultados 2023:         {len(RESULTADO_2023)} municipios")
print(f"  Hoja 4 – Legisladores Prov.:      {len(DIPUTADOS_PROV)+len(SENADORES_PROV)} ({len(DIPUTADOS_PROV)} dip + {len(SENADORES_PROV)} sen)")
print(f"  Hoja 5 – Legisladores Nac.:       {len(DIPUTADOS_NAC)+len(SENADORES_NAC)} ({len(DIPUTADOS_NAC)} dip + {len(SENADORES_NAC)} sen)")
print(f"  Hoja 6 – Concejales 2023:         {_h6_count} concejales ({len(CONCEJALES_2023)} municipios)")
print(f"  Hoja 7 – Resumen ejecutivo")
