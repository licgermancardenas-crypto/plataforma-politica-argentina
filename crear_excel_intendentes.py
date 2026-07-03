#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# DATOS: 135 intendentes provincia de Buenos Aires (mandato 2023-2027)
# Fuente: La Tecla / Infocielo - elecciones octubre 2023
# ─────────────────────────────────────────────────────────────

INTENDENTES = [
    ("Adolfo Alsina",           "Javier Andrés",              "Unión por la Patria"),
    ("Adolfo Gonzales Chaves",  "Miriam Gómez",               "Juntos por el Cambio"),
    ("Alberti",                 "Germán Lago",                "Unión por la Patria"),
    ("Almirante Brown",         "Mariano Cascallares",        "Unión por la Patria"),
    ("Arrecifes",               "Fernando Bouvier",           "Juntos por el Cambio"),
    ("Avellaneda",              "Jorge Ferraresi",            "Unión por la Patria"),
    ("Ayacucho",                "Emilio Cordonnier",          "Juntos por el Cambio"),
    ("Azul",                    "Nelson Sombra",              "Unión por la Patria"),
    ("Bahía Blanca",            "Federico Susbielles",        "Unión por la Patria"),
    ("Balcarce",                "Esteban Reino",              "Juntos por el Cambio"),
    ("Baradero",                "Esteban Sanzio",             "Unión por la Patria"),
    ("Benito Juárez",           "Julio Marini",               "Unión por la Patria"),
    ("Berazategui",             "Juan José Mussi",            "Unión por la Patria"),
    ("Berisso",                 "Fabián Cagliardi",           "Unión por la Patria"),
    ("Bolívar",                 "Marcos Pisano",              "Unión por la Patria"),
    ("Bragado",                 "Sergio Barenghi",            "Unión por la Patria"),
    ("Brandsen",                "Fernando Raitelli",          "Unión por la Patria"),
    ("Campana",                 "Sebastián Abella",           "Juntos por el Cambio"),
    ("Cañuelas",                "Marisa Fassi",               "Unión por la Patria"),
    ("Capitán Sarmiento",       "Fernanda Astorino",          "Juntos por el Cambio"),
    ("Carlos Casares",          "Daniel Stadnik",             "Unión por la Patria"),
    ("Carlos Tejedor",          "María Celia Gianini",        "Unión por la Patria"),
    ("Carmen de Areco",         "Iván Villagrán",             "Unión por la Patria"),
    ("Castelli",                "Francisco Echarren",         "Unión por la Patria"),
    ("Chacabuco",               "Rubén Darío Golía",          "Unión por la Patria"),
    ("Chascomús",               "Javier Gastón",              "Unión por la Patria"),
    ("Chivilcoy",               "Guillermo Britos",           "Partido Chivilcoy (vecinal)"),
    ("Colón",                   "Waldemar Giordano",          "Unión por la Patria"),
    ("Coronel Dorrego",         "Juan C. Chalde",             "Juntos por el Cambio"),
    ("Coronel Pringles",        "Lisandro Matzkin",           "Juntos por el Cambio"),
    ("Coronel Rosales",         "Rodrigo Aristimuño",        "Unión por la Patria"),
    ("Coronel Suárez",          "Ricardo Moccero",            "Unión por la Patria"),
    ("Daireaux",                "Alejandro Acerbo",           "Unión por la Patria"),
    ("Dolores",                 "Juan Pablo García",          "Unión por la Patria"),
    ("Ensenada",                "Mario Secco",                "Unión por la Patria"),
    ("Escobar",                 "Ariel Sujarchuk",            "Unión por la Patria"),
    ("Esteban Echeverría",      "Fernando Gray",              "Unión por la Patria"),
    ("Exaltación de la Cruz",   "Diego Nanni",                "Unión por la Patria"),
    ("Ezeiza",                  "Gastón Granados",            "Unión por la Patria"),
    ("Florencio Varela",        "Andrés Watson",              "Unión por la Patria"),
    ("Florentino Ameghino",     "Nahuel Mittelbach",          "Juntos por el Cambio"),
    ("General Alvarado",        "Sebastián Ianantuony",       "Unión por la Patria"),
    ("General Alvear",          "Ramón José Capra",           "Juntos por el Cambio"),
    ("General Arenales",        "Erica Revilla",              "Juntos por el Cambio"),
    ("General Belgrano",        "Osvaldo Dinápoli",           "Juntos por el Cambio"),
    ("General Guido",           "Carlos Rocha",               "Unión por la Patria"),
    ("General La Madrid",       "Martín Randazzo",            "Juntos por el Cambio"),
    ("General Las Heras",       "Javier Osuna",               "Unión por la Patria"),
    ("General Lavalle",         "Nahuel Guardia",             "Juntos por el Cambio"),
    ("General Madariaga",       "Carlos Santoro",             "Juntos por el Cambio"),
    ("General Paz",             "Juan Manuel Alvarez",        "Unión por la Patria"),
    ("General Pinto",           "Alfredo 'Freddy' Zavatarelli", "Unión por la Patria"),
    ("General Pueyrredon",      "Guillermo Montenegro",       "Juntos por el Cambio"),
    ("General Rodríguez",       "Mauro Santiago García",      "Unión por la Patria"),
    ("General San Martín",      "Fernando Moreira",           "Unión por la Patria"),
    ("General Viamonte",        "Franco Flexas",              "Juntos por el Cambio"),
    ("General Villegas",        "Gilberto Alegre",            "Juntos por el Cambio"),
    ("Guaminí",                 "José Nobre Ferreira",        "Unión por la Patria"),
    ("Hipólito Yrigoyen",       "Ignacio Pugnaloni",          "Unión por la Patria"),
    ("Hurlingham",              "Damián Selci",               "Unión por la Patria"),
    ("Ituzaingó",               "Pablo Descalzo",             "Unión por la Patria"),
    ("José C. Paz",             "Mario Ishii",                "Unión por la Patria"),
    ("Junín",                   "Pablo Petrecca",             "Juntos por el Cambio"),
    ("La Costa",                "Juan De Jesús",              "Unión por la Patria"),
    ("La Matanza",              "Fernando Espinoza",          "Unión por la Patria"),
    ("La Plata",                "Julio Alak",                 "Unión por la Patria"),
    ("Lanús",                   "Julián Álvarez",             "Unión por la Patria"),
    ("Laprida",                 "Alfredo 'Pichi' Fisher",     "Unión por la Patria"),
    ("Las Flores",              "Alberto Gelené",             "Unión por la Patria"),
    ("Leandro N. Alem",         "Carlos Ferraris",            "Unión por la Patria"),
    ("Lezama",                  "Arnaldo Harispe",            "Juntos por el Cambio"),
    ("Lincoln",                 "Salvador Serenal",           "Juntos por el Cambio"),
    ("Lobería",                 "Pablo Barrena",              "Juntos por el Cambio"),
    ("Lobos",                   "Jorge Etcheverry",           "Juntos por el Cambio"),
    ("Lomas de Zamora",         "Federico Otermín",           "Unión por la Patria"),
    ("Luján",                   "Leonardo Boto",              "Unión por la Patria"),
    ("Magdalena",               "Lisandro Hourcade",          "Juntos por el Cambio"),
    ("Maipú",                   "Matías Rappallini",          "Juntos por el Cambio"),
    ("Malvinas Argentinas",     "Leonardo Nardini",           "Unión por la Patria"),
    ("Mar Chiquita",            "Walter Wischnivetzky",       "Unión por la Patria"),
    ("Marcos Paz",              "Ricardo Curutchet",          "Unión por la Patria"),
    ("Mercedes",                "Juan Ignacio Ustarroz",      "Unión por la Patria"),
    ("Merlo",                   "Gustavo Menéndez",           "Unión por la Patria"),
    ("Monte",                   "José M. Castro",             "Juntos por el Cambio"),
    ("Monte Hermoso",           "Hernán Arranz",              "Unión por la Patria"),
    ("Moreno",                  "Mariel Fernández",           "Unión por la Patria"),
    ("Morón",                   "Lucas Ghi",                  "Unión por la Patria"),
    ("Navarro",                 "Facundo Diz",                "Unión por la Patria"),
    ("Necochea",                "Arturo Rojas",               "Nuevo Necochea (vecinal)"),
    ("Nueve de Julio",          "María Josefina Gentile",     "Juntos por el Cambio"),
    ("Olavarría",               "Maximiliano Wesner",         "Unión por la Patria"),
    ("Patagones",               "Ricardo Marino",             "Unión por la Patria"),
    ("Pehuajó",                 "Pablo Zurro",                "Unión por la Patria"),
    ("Pellegrini",              "Sofía Gambier",              "Juntos por el Cambio"),
    ("Pergamino",               "Javier Martínez",            "Juntos por el Cambio"),
    ("Pila",                    "Gustavo S. Walker",          "Unión por la Patria"),
    ("Pilar",                   "Federico Achával",           "Unión por la Patria"),
    ("Pinamar",                 "Juan M. Ibarguren",          "Juntos por el Cambio"),
    ("Presidente Perón",        "Blanca Cantero",             "Unión por la Patria"),
    ("Puán",                    "Diego Reyes",                "Juntos por el Cambio"),
    ("Punta Indio",             "David Angueira",             "Unión por la Patria"),
    ("Quilmes",                 "Mayra Mendoza",              "Unión por la Patria"),
    ("Ramallo",                 "Mauro Poletti",              "Unión por la Patria"),
    ("Rauch",                   "M. Suescun",                 "Juntos por el Cambio"),
    ("Rivadavia",               "Juan Alberto Martínez",      "Unión por la Patria"),
    ("Rojas",                   "Román Bouvier",              "Juntos por el Cambio"),
    ("Roque Pérez",             "Maximiliano Sciaini",        "Unión por la Patria"),
    ("Saavedra",                "Matías Nebot",               "Todos x Saavedra (vecinal)"),
    ("Saladillo",               "José Luis Salomón",          "Juntos por el Cambio"),
    ("Salliqueló",              "Ariel Succurro",             "Unión por la Patria"),
    ("Salto",                   "Ricardo 'Cura' Alessandro",  "Unión por la Patria"),
    ("San Andrés de Giles",     "Miguel Gesualdi",            "Unión por la Patria"),
    ("San Antonio de Areco",    "Francisco Ratto",            "Juntos por el Cambio"),
    ("San Cayetano",            "M. A. Gargaglione",          "Juntos por el Cambio"),
    ("San Fernando",            "Juan Andreotti",             "Unión por la Patria"),
    ("San Isidro",              "Ramón Lanús",                "Juntos por el Cambio"),
    ("San Miguel",              "Jaime Méndez",               "Juntos por el Cambio"),
    ("San Nicolás",             "Santiago Passaglia",         "Juntos por el Cambio"),
    ("San Pedro",               "Cecilio Salazar",            "Unión por la Patria"),
    ("San Vicente",             "Nicolás Mantegazza",         "Unión por la Patria"),
    ("Suipacha",                "Juan Luis Mancini",          "Unión por la Patria"),
    ("Tandil",                  "Miguel Lunghi",              "Juntos por el Cambio"),
    ("Tapalqué",                "Gustavo Cocconi",            "Unión por la Patria"),
    ("Tigre",                   "Julio Zamora",               "Unión por la Patria"),
    ("Tordillo",                "Héctor Olivera",             "Unión por la Patria"),
    ("Tornquist",               "Sergio Bordoni",             "Unión por la Patria"),
    ("Trenque Lauquen",         "Francisco Recoulat",         "Juntos por el Cambio"),
    ("Tres Arroyos",            "Pablo Garate",               "Unión por la Patria"),
    ("Tres de Febrero",         "Diego Valenzuela",           "Juntos por el Cambio"),
    ("Tres Lomas",              "Luciano Spinolo Sayago",     "Juntos por el Cambio"),
    ("Veinticinco de Mayo",     "Ramiro Egüen",               "Juntos por el Cambio"),
    ("Vicente López",           "Soledad Martínez",           "Juntos por el Cambio"),
    ("Villa Gesell",            "Gustavo Barrera",            "Unión por la Patria"),
    ("Villarino",               "Carlos Bevilacqua",          "Agrupación x Villarino (vecinal)"),
    ("Zárate",                  "Marcelo Matzkin",            "Juntos por el Cambio"),
]

# ─────────────────────────────────────────────────────────────
# SECRETARIOS (relevados de sitios oficiales municipales)
# ─────────────────────────────────────────────────────────────

SECRETARIOS = {
    "La Matanza": [
        ("Silvia Francese",          "Secretaria General de Gobierno"),
        ("Claudio Aulicino",         "Secretario de Economía y Hacienda"),
        ("Héctor Turquié",           "Secretario de Obras y Servicios Públicos"),
        ("Alejandro Collia",         "Secretario de Salud Pública"),
        ("Débora Giorgi",            "Secretaria de Producción"),
        ("Silvina Gvirtz",           "Secretaria de Políticas Educativas"),
        ("Natalia Lucchetti",        "Secretaria General Privada"),
        ("Liliana Hendel",           "Secretaria de Mujeres, Políticas de Género y Diversidades"),
        ("Mirta Hembert",            "Secretaria de Desarrollo Social"),
        ("Eduardo Barcat",           "Secretario General de Protección Ciudadana"),
        ("Miguel Saredi",            "Secretario de Planificación Operativa"),
        ("Antonio Colicigno",        "Secretario Gral. de Coordinación de Gestión e Innovación"),
        ("Jorge Di Santo",           "Secretario General de Espacios y Servicios Públicos"),
        ("Belén Pagni",              "Secretaria de Juventudes"),
        ("Rocío Giussi",             "Secretaria de Control Comunal"),
        ("Jorgelina Bertoni",        "Secretaria de Deportes y Recreación"),
        ("Javier Fernández Castro",  "Secretario de Hábitat"),
        ("Marisa Guerin",            "Secretaria de la Tercera Edad"),
        ("Ricardo Rolleri",          "Secretario de Desarrollo Estratégico"),
        ("Claudio Lentini",          "Jefe de Gabinete"),
        ("Nicolás Fusca",            "Vicejefe de Gabinete"),
    ],
    "Florencio Varela": [
        ("Christian Rodríguez",        "Secretaría General de Gestión Pública"),
        ("Andrea Padial",              "Secretaría de Coordinación Técnico Jurídica"),
        ("Estefanía Gabriela Nieva",   "Secretaría de Cultura, Deportes y Recreación"),
        ("Laura Mabel Vivas",          "Secretaría de Desarrollo Social"),
        ("Andrea Fabiana Digiobani",   "Secretaría de Educación"),
        ("Gisella Elizabeth Primus",   "Secretaría de Hacienda"),
        ("Daniel Alejandro Dono Leidi","Secretaría de Gobierno"),
        ("Ricardo Sebastián de la Fuente","Secretaría de Industria y Desarrollo Productivo"),
        ("Adriana Elda Schulz",        "Secretaría Legal y Técnica"),
        ("Diego Hernán Trejo",         "Secretaría de Obras, Servicios Públicos, Ambiente y Planif. Urbana"),
        ("Alberto Oscar Núñez",        "Secretaría de Relaciones Institucionales"),
        ("Adriana Estela Alonso",      "Secretaría de Salud"),
    ],
    "Lomas de Zamora": [
        ("Martín Choren",      "Jefe de Gabinete"),
        ("María Sol Tischik",  "Secretaria General"),
        ("Víctor Matassi",     "Secretario Técnico Legal Administrativo"),
        ("Oscar Ferreyra",     "Secretario de Gobierno"),
        ("Marcelo D'Angelo",   "Secretario de Finanzas"),
        ("Emiliano Piergiovanni", "Secretario de Obras Públicas"),
        ("Marina Lesci",       "Secretaria de Relaciones con la Comunidad"),
        ("Maximiliano Tonani", "Secretario de Seguridad"),
        ("Lucas Modarelli",    "Secretario de Desarrollo Social"),
        ("Mariano Ortega",     "Secretario de Salud"),
        ("Matías Gasparrini",  "Secretario de Cultura y Educación"),
        ("Pedro Procopio",     "Secretario de Ambiente"),
        ("Luciana Angueira",   "Área de Género y Mujeres"),
        ("Victoria Bourio",    "Secretaría Privada"),
    ],
    "Morón": [
        ("Estefanía Franco",   "Jefa de Gabinete"),
        ("Guido Napolitano",   "Secretario de Economía y Finanzas"),
        ("Damián Cardoso",     "Secretario de Seguridad Ciudadana"),
        ("Claudio Román",      "Secretario de Control Comunal"),
        ("Santiago Muñiz",     "Secretario de Desarrollo Productivo"),
        ("José María Ghi",     "Secretario de Educación y Desarrollo de la Comunidad"),
        ("Guillermo Pascuero", "Secretario de Planificación Estratégica"),
        ("Jacobo Netel",       "Secretario de Salud"),
        ("Eugenia Navarro",    "Secretaria de Desarrollo Local, Empleo y Economía Social"),
        ("Laura De Peri",      "Secretaria de Mujeres, Géneros y Diversidad"),
        ("Osvaldo Carballo",   "Secretario de Obras y Servicios Públicos"),
        ("Oscar Conde",        "Secretario de Tránsito y Transporte"),
    ],
    "Esteban Echeverría": [
        ("Miguel Ángel Urchipia",  "Jefe de Gabinete"),
        ("Valeria Alejandra Bellizzi", "Secretaria de Gobierno"),
        ("Walter Saracco",         "Secretario de Hacienda"),
        ("Sergio Benet",           "Secretario de Obras Públicas"),
        ("Diego Fetissoff",        "Secretario de Servicios Públicos"),
        ("Emiliano Valentino",     "Secretario de Desarrollo Social"),
        ("Gabriel Emiliano Ive",   "Secretario de Salud"),
        ("Gabriel Villegas",       "Secretario de Seguridad Comunal"),
        ("Alejandro Correa",       "Secretario de Cultura"),
        ("Marcelo Guillermo González", "Secretario de Tierras y Vivienda"),
        ("Roberto Ariel Devoto",   "Secretario de Desarrollo Local, Políticas Públicas y Modernización"),
    ],
    "Tres de Febrero": [
        ("Fernando Ramos",     "Secretario de Gobierno"),
        ("Federico Calvo",     "Secretario de Hacienda"),
        ("Alberto Nazer",      "Secretario de Obras Públicas"),
        ("Rodrigo Aybar",      "Secretario de Ambiente y Servicios Públicos"),
        ("Cecilia Pampuro",    "Secretaria de Salud"),
        ("Inés García",        "Secretaria de Desarrollo Humano"),
        ("Mercedes Sanguineti","Secretaria de Educación y Cultura"),
        ("Bautista Pino",      "Secretario de Asuntos Estratégicos y Coordinación"),
        ("Daniela Ramos",      "Secretaria de Trabajo y Producción"),
        ("Pablo Casna",        "Secretario de Atención al Vecino"),
        ("Juan Marchese",      "Secretario de Seguridad"),
        ("Pablo Fuentes",      "Secretario de Deportes"),
    ],
    "San Isidro": [
        ("Manuela López Menéndez", "Jefa de Gabinete"),
        ("Santiago Dondo",         "Secretario General"),
        ("José Sánchez Sorondo",   "Secretario de Gobierno"),
        ("Enrique Rodríguez Varela","Secretario de Seguridad"),
        ("Juan Sanguinetti",       "Secretario de Hacienda"),
        ("Mercedes Sanguineti",    "Secretaria de Educación, Cultura y Trabajo"),
        ("Laura Leali",            "Secretaria de Legal y Técnica"),
        ("Clara Sanguinetti",      "Secretaria de Ambiente y Espacio Público"),
        ("Santiago Mayorga",       "Subsecretario de Estrategia, Comunicación y Prensa"),
        ("Alec Lucena",            "Subsecretario de Innovación"),
        ("Diego Martín Cortina",   "Agencia de Recaudación"),
        ("Manuel Abella Nazar",    "Subsecretario de Gestión y Participación"),
        ("Lucas Prieto Schorr",    "Subsecretario de Relaciones con la Comunidad"),
        ("Ignacio Videla",         "Subsecretario de Salud"),
        ("Javier Jaureguiberry",   "Subsecretario de Producción y Empleo"),
        ("Florencia Aguilar",      "Área de Planeamiento Urbano"),
    ],
    "Lanús": [
        ("Nadia Burgos",        "Jefa de Gabinete"),
        ("Leandro 'Lalo' Decuzzi", "Secretario de Gobierno"),
        ("Laura Alonso",        "Secretaria de Desarrollo Social"),
        ("Mauro Iezzi",         "Secretario de Espacio Público"),
        ("Emilia Aristei",      "Secretaria de Obras Públicas"),
        ("Natalia Gradaschi",   "Secretaria de Educación"),
        ("Sebastián Castillo",  "Secretario de Seguridad"),
        ("Nicolás Russo",       "Secretario de Producción y Comercio"),
        ("Alejo Di Carlo",      "Secretario de Cultura y Deportes"),
        ("Maru Decuzzi",        "Secretaria de Prensa y Comunicaciones"),
    ],
    "Bahía Blanca": [
        ("Luis Alberto Calderaro", "Jefe de Gabinete"),
        ("Florencia Molini",       "Secretaria de Gobierno"),
        ("Carlos Fernando De Vadillo", "Secretario de Economía"),
        ("Gustavo Guillermo Trankels", "Secretario de Obras y Servicios Públicos"),
        ("José María Zingoni Segatori", "Secretario de Planeamiento Urbano"),
        ("Ignacio Caspe",          "Secretario de Coordinación de Delegaciones"),
        ("Federico Aníbal Bugatti","Secretario de Salud"),
        ("Romina Pires",           "Secretaria de Políticas Sociales y Fortalecimiento Comunitario"),
    ],
    "Quilmes": [
        ("Joaquín Desmery",    "Jefe de Gabinete"),
        ("Florencia Esteche",  "Secretaria de Gobierno"),
        ("Nicolás Mellino",    "Secretario de Deportes"),
        ("Evangelina Ramírez", "Secretaria de Educación y Culturas"),
    ],
    "Pilar": [
        ("Santiago Laurent",   "Secretario de Gobierno y Desarrollo Comunitario"),
        ("Darío Genua",        "Secretario de Economía y Hacienda"),
        ("Soledad Peralta",    "Secretaria General"),
        ("Federico Kruse",     "Secretario de Seguridad"),
        ("Paula González",     "Secretaria de Desarrollo Social"),
        ("Guido Bordachar",    "Secretario de Obras y Servicios Públicos"),
        ("Mirta Ortega Sanz",  "Secretaria de Salud"),
        ("Eduardo Martignoni", "Secretaria de Prensa"),
    ],
    "General Pueyrredon": [
        ("Mauro Martinelli",   "Secretario de Hacienda y Legal y Técnica"),
        ("Rodrigo Goncalvez",  "Secretario de Seguridad"),
        ("Fernando Rizzi",     "Secretario de Educación"),
        ("Viviana Bernabei",   "Secretaria de Salud"),
    ],
    "Almirante Brown": [
        ("Juan Fabiani",          "Secretario de Gobierno"),
        ("Alan Grin",             "Secretario General"),
        ("Miguel Faienza",        "Secretario de Economía y Hacienda"),
        ("Walter Gómez",          "Secretario de Salud"),
        ("Bárbara Miñán",         "Secretaria de Desarrollo, Seguridad Social y Derechos Humanos"),
        ("Atilio Gari",           "Secretario de Gestión Descentralizada"),
        ("Paula Eichel",          "Secretaria de Prevención y Seguridad Ciudadana"),
        ("Federico Sassone",      "Secretario de Producción, Empleo y Formación Profesional"),
        ("Sergio Pianciola",      "Secretario de Educación, Ciencia y Tecnología"),
        ("Ignacio Villaronga",    "Secretario de Política Ambiental y Hábitat"),
        ("Fernando Lorenzo",      "Secretario de Infraestructura, Planificación y Servicios Públicos"),
    ],
    "Avellaneda": [
        ("Magdalena Sierra",         "Jefa de Gabinete"),
        ("Federico Bonaldi",         "Secretario de Cultura"),
        ("Romina Barreiro",          "Secretaria de Desarrollo Social"),
        ("Fabián Monzón",            "Secretario de Gobierno"),
        ("Mónica Ghirelli",          "Secretaria de Desarrollo Territorial y Hábitat"),
        ("Alejo Chornobroff",        "Secretario de Seguridad"),
        ("Claudio Yacoy",            "Secretario de Derechos Humanos"),
        ("Gastón Seillan",           "Secretario de Obras y Servicios Públicos"),
        ("Guillermo Pesce",          "Secretario de Planeamiento"),
        ("Virginia González Algañaraz", "Secretaria de Salud"),
        ("María Gabriela Sierra",    "Secretaria de Hacienda y Administración"),
        ("Carlos Lombardo",          "Secretario de Producción, Comercio y Ambiente"),
        ("Sergio Ceballos",          "Secretario Legal y Técnico"),
        ("Sebastián Vidal",          "Secretario de Deportes"),
    ],
    "Berisso": [
        ("Ramiro Crilchuk",    "Jefe de Gabinete"),
        ("Gabriel Pérez",      "Secretaría Privada"),
        ("Aldana Iovanovich",  "Secretaria de Gobierno"),
        ("Gabriel Bruno",      "Secretario de Economía"),
        ("Roberto Alonso",     "Secretario de Producción"),
        ("Rita Hernández",     "Secretaria de Salud"),
        ("Lucas Spivak",       "Secretario de Promoción Social"),
    ],
    "Ensenada": [
        ("María Alejandra Sabio",    "Secretaria de Gobierno"),
        ("Marcos Omentari",          "Secretario de Inspección y Control Urbano"),
        ("Edgardo Reyes",            "Secretario de Servicios Públicos"),
        ("José Alberto Núñez",       "Secretario de Obras Públicas"),
        ("Rocío Basso",              "Secretaria de Hacienda y Producción"),
        ("María Celina Ferella",     "Secretaria Privada"),
        ("Martín Slobodian",         "Secretario de Seguridad y Justicia"),
        ("Agustín Duscovich",        "Secretario de Gestión Pública"),
        ("Juan Manuel López Ortega", "Secretario de Salud y Ambiente"),
    ],
    "Ezeiza": [
        ("Pablo Trombetta",    "Jefatura de Gabinete Interino"),
        ("Edgardo Amarilla",   "Jefatura de Gabinete de Asesores"),
        ("Gustavo Boccaccio",  "Ambiente y Construcción Comunitaria"),
        ("Alejandro Carballo", "Obras Públicas"),
        ("Gabriel Robles",     "Contaduría"),
        ("Dulce Granados",     "Cultura"),
        ("Laura Arnal",        "Desarrollo y Trabajo para la Comunidad"),
        ("José Danelotti",     "Economía y Hacienda"),
        ("Mario Calvella",     "Educación, Deportes y Turismo"),
        ("Miriam San Martín",  "Gobierno"),
        ("Juan Eder",          "Ingresos Municipales"),
        ("Héctor Canzani",     "Inspección General"),
        ("Emmanuel Amarilla",  "Juventudes"),
        ("Jorge Laurino",      "Legal y Técnica"),
        ("Mariano Funes",      "Modernización"),
        ("Natalia Sena",       "Mujer, Políticas de Género y Diversidades"),
        ("Mónica Marchetti",   "Personal"),
        ("Juan Arrigoni",      "Planeamiento, Catastro y Obras Particulares"),
        ("Carolina Basso",     "Protección y Bienestar Animal"),
        ("Valeria Amigo",      "Salud"),
        ("Marisol Granados",   "Secretaría Privada"),
        ("Hugo Matzkin",       "Seguridad"),
        ("Esteban Carrizo",    "Servicios"),
        ("Walter Álvarez",     "Trabajo"),
    ],
    "General San Martín": [
        ("Alejandro Tsolis",   "Jefe de Gabinete"),
        ("Luciano Miranda",    "Secretario de Salud"),
        ("Marcela Ferri",      "Secretaria de Mujeres, Géneros y Niñeces"),
        ("Marcos Vilaplana",   "Secretario Legal y Técnico"),
        ("Andrés Alonso",      "Secretario de Obras Públicas y Servicios"),
    ],
    "Hurlingham": [
        ("Florencia Lampreabe","Jefa de Gabinete"),
        ("Miguel Quintero",    "Vicejefe de Gabinete"),
        ("Lorena Fabis",       "Secretaria de Gobierno"),
        ("Facundo Pérez",      "Secretario de Hacienda"),
        ("Elena Cerviño",      "Secretaria de Desarrollo Social"),
        ("Daniela Campisi",    "Secretaria de Cultura"),
        ("Nicolás Coliqueo",   "Secretario de Salud"),
        ("Ariel Simoni",       "Secretario de Comercio"),
        ("Facundo Cadavid",    "Secretario de Infraestructura y Obras Públicas"),
        ("Adrián Eslaiman",    "Secretario de Educación"),
        ("Luca Pratti",        "Secretario de Seguridad"),
    ],
    "Ituzaingó": [
        ("Pablo Andrés Piana",    "Jefe de Gabinete"),
        ("Mauro Lavagnino",       "Secretario de Gobierno"),
        ("Fabián Ariel Basílico", "Secretario de Salud"),
        ("Sandra Rey",            "Secretaria de Innovación, Transparencia y Asuntos Estratégicos"),
        ("Marcelo Martinelli",    "Secretario de Economía y Hacienda"),
        ("Natalia Lazzeri",       "Secretaria de Desarrollo Productivo"),
        ("Antonio Rosendo",       "Secretario de Seguridad"),
        ("Martín Rossi",          "Secretario de Planificación, Desarrollo Urbano y Ambiente"),
        ("Juan Cruz Descalzo",    "Secretario de Servicios Públicos"),
        ("Ayelén Estévez",        "Secretaria de Cultura y Educación"),
        ("Lucas García",          "Secretario de Deportes"),
        ("Nahuel Segovia",        "Secretario de Desarrollo Humano y Relaciones con la Comunidad"),
        ("Rita Georgina Sale",    "Secretaria de Educación Técnica y Superior"),
        ("Silvina Morelli",       "Subsecretaria de Imagen y Comunicación"),
    ],
    "José C. Paz": [
        ("Gastón Yáñez",          "Jefe de Gabinete"),
        ("Juan José Pérez",       "Secretario de Gobierno"),
        ("Juan Pablo Mansilla",   "Secretario de Acción Directa"),
        ("Juan Carlos Denuchi",   "Secretario de Defensa al Consumidor"),
        ("Nélida Barboza",        "Secretaria de Desarrollo Social"),
        ("Roberto Caggiano",      "Secretario de Obras Públicas"),
        ("Humberto Fernández",    "Secretario de Economía y Hacienda"),
        ("Fernando Cabrera",      "Secretario de Seguridad"),
    ],
    "La Plata": [
        ("Carlos Eduardo Bonicatto",  "Jefe de Gabinete"),
        ("Aníbal Norberto Gómez",     "Secretario General"),
        ("Guillermo Cara",            "Secretario de Gobierno"),
        ("Luis Federico Arias",       "Secretario de Coordinación"),
        ("Sergio Resa",               "Secretario de Planeamiento, Obras y Servicios Públicos"),
        ("Marcelo Darío Giampaoli",   "Secretario de Hacienda y Finanzas"),
        ("Guillermo Martín Escudero", "Secretario de Ambiente"),
        ("Paula Lambertini",          "Secretaria de Educación"),
        ("María Soledad Fernández",   "Secretaria de Salud"),
        ("Ana Amelia Negrete",        "Secretaria de Cultura"),
        ("Silvina Perugino",          "Secretaria de Mujeres y Diversidad"),
        ("Marina Mongiardino",        "Secretaria de Justicia"),
        ("Diego Mario Abel Pepe",     "Secretario de Seguridad"),
        ("Víctor Hortel",             "Secretario de Control Urbano"),
        ("Mercedes La Giosa",         "Secretaria de Producción"),
        ("Gastón Castagnetto",        "Secretario de Relaciones Institucionales"),
        ("Tomás Barbier",             "Secretario de Modernización"),
        ("Claudia Azucena Gallardo",  "Secretaria de Economía Popular"),
        ("Guillermo Federico Comadira","Secretario Legal y Técnico"),
        ("Nicolás Carvalho",          "Secretario de Desarrollo Social"),
    ],
    "Luján": [
        ("Javier Legorburu",     "Secretario de Servicios Públicos"),
        ("Mariana Girón",        "Secretaria de Salud"),
        ("Silvio Martini",       "Secretario de Gobierno y Participación Ciudadana"),
        ("Nicolás Capelli",      "Secretario de Culturas y Turismo"),
        ("Federico Vanin",       "Secretario de Desarrollo Humano"),
        ("Virginia Martinelli",  "Secretaria Legal y Técnica"),
    ],
    "Merlo": [
        ("Gustavo Soos",        "Jefe de Gabinete"),
        ("Carlos Dobler",       "Secretario de Gobierno"),
        ("Mauricio Canosa",     "Secretario de Delegaciones"),
        ("Lucas Scarcella",     "Secretario de Desarrollo e Integración Social"),
        ("Walter Videtta",      "Secretario de Salud"),
    ],
    "Moreno": [
        ("Alberto Conca",         "Secretario General"),
        ("Nahuel Berguier",       "Secretario de Gobierno y Justicia"),
        ("Mariela Bien",          "Secretaria de Economía"),
        ("María Gimenez",         "Secretaria de Obras y Servicios Públicos"),
        ("Juan Varani",           "Secretario de Salud"),
        ("Cynthia Muñoz",         "Secretaria de Educación, Cultura y Deporte"),
        ("Mirna Coronel",         "Secretaria de Mujeres, Géneros y Diversidades"),
        ("Daiana Anadón",         "Secretaria de Ambiente y Desarrollo Sostenible"),
        ("Ismael Castro",         "Secretario de Seguridad"),
        ("Martín Fraiz",          "Secretario de Tránsito y Transporte"),
        ("Martín Linares",        "Secretario de Comunicación y Relaciones Internacionales"),
        ("Marcela Diaz",          "Secretaria Privada"),
        ("Lis Díaz",              "Secretaria de Desarrollo Comunitario"),
    ],
    "Olavarría": [
        ("Mercedes Landívar",      "Jefa de Gabinete"),
        ("Cristian Daniel Delpiani","Secretario de Desarrollo de la Comunidad"),
        ("Eugenia Bezzoni",        "Secretaria de Economía y Hacienda"),
        ("Orfel Fariña",           "Secretario de Obras Públicas, Infraestructura y Servicios Públicos"),
        ("Gastón Sarachu",         "Secretario de Desarrollo Económico y Productivo"),
    ],
    "Pergamino": [
        ("Carlos Damián Pérez",    "Jefe de Gabinete"),
        ("Juan Manuel Rico Zini",  "Secretario de Gobierno"),
        ("Matías Villeta",         "Secretario de Salud"),
        ("Sergio Tressens",        "Secretario de Finanzas"),
        ("José Salauati",          "Secretario de Desarrollo Urbano"),
        ("Sergio Pizarro",         "Secretario de Producción"),
        ("Karim Dib",              "Secretario de Seguridad y Tránsito"),
    ],
    "San Fernando": [
        ("Ignacio Pérez",     "Secretario de Gobierno"),
        ("Germán Roldán",     "Secretario de Economía"),
        ("Néstor Torchia",    "Secretario de Cultura y Turismo"),
    ],
    "San Miguel": [
        ("Joaquín Estrada",    "Jefe de Gabinete"),
        ("Francisco Nigro",    "Secretario de Gobierno"),
        ("Marcelo Conzi",      "Secretario de Economía y Finanzas"),
        ("Andrés Lagalaye",    "Secretario de Educación y Trabajo"),
        ("Yesica Grillo",      "Secretaria de Infancia y Familia"),
        ("Federico Randle",    "Secretario de Obras Públicas y Espacios"),
        ("Milagros Richards",  "Secretaria de Planificación y Desarrollo Urbano"),
        ("Luis Costán",        "Secretario de Salud"),
        ("Héctor Calvente",    "Secretario de Seguridad"),
        ("Franco Ortíz",       "Secretario de Comunicación y Deportes"),
    ],
    "Tigre": [
        ("Fernando Lauría",     "Secretario General y de Economía"),
        ("Mario Zamora",        "Secretario de Gobierno"),
        ("Pedro Heyde",         "Secretario de Servicios Públicos"),
        ("Fernando Abramzon",   "Secretario de Salud"),
        ("Gisela Zamora",       "Secretaria de Desarrollo Social y Políticas de Inclusión"),
        ("Emiliano Mansilla",   "Secretario de Desarrollo Económico y Relaciones con la Comunidad"),
        ("Eduardo Feijoo",      "Secretario de Protección Ciudadana"),
        ("Florencia Mosqueda",  "Secretaria de Turismo y Medio Ambiente"),
        ("Galdys Pollán",       "Secretaria de Planificación y Desarrollo del Espacio Territorial"),
        ("Carolina Álvarez Eguileta", "Agencia de Educación"),
        ("Marilina Silva",      "Agencia de Cultura"),
        ("Natalia Gómez",       "Agencia de Deporte"),
    ],
    "Trenque Lauquen": [
        ("Alfredo Luis Zambiasio", "Secretario de Hacienda"),
        ("Martín Ignacio Borrazas","Secretario de Gobierno"),
        ("Gustavo Andrés Marchabalo","Secretario Legal y Técnico"),
        ("Manuel Silva Muñoz",    "Secretario de Obras y Servicios Públicos"),
        ("Yanina López",          "Secretaria de Salud"),
        ("Germán Lauro",          "Subsecretario de Desarrollo Económico y Productivo"),
    ],
    "Vicente López": [
        ("Facundo Carrillo",           "Jefe de Gabinete"),
        ("Carlos Alberto Granovsky",   "Secretario de Comunicación"),
        ("Luciana Blasco",             "Secretaria de Cultura"),
        ("Cristian Darío Muscillo",    "Secretario de Deportes"),
        ("Antonella Regina Del Prete", "Secretaria de Desarrollo Social"),
        ("Rocío Soledad Fontana",      "Secretaria de Educación y Empleo"),
        ("Enio Vittorini",             "Secretario de Gobierno, Legal y Técnica"),
        ("Miguel Álvarez",             "Secretario de Hacienda y Finanzas"),
        ("Mariano Carlos Botto",       "Secretario de Planeamiento y Obras Públicas"),
        ("Ángeles Martínez",           "Secretaria de Salud"),
        ("Santiago de Jesús",          "Secretario de Seguridad"),
        ("Leandro Jorge Martín",       "Secretario de Servicios Públicos"),
        ("Juan Pablo Fittipaldi",      "Secretario de Transformación Digital"),
    ],
    "Zárate": [
        ("Alejandra Elena Lozano",     "Secretaria General"),
        ("Juan Bautista Fernández",    "Secretario de Desarrollo Económico"),
        ("Cristian Mazzola",           "Secretario de Hacienda y Finanzas"),
        ("Juan Manuel Iglesias",       "Secretario de Seguridad Ciudadana"),
        ("Sergio Ricardo Agostinelli", "Secretario de Servicios Públicos"),
        ("Ivan Rodrigo Gómez Gerez",   "Secretario de Desarrollo Humano y Promoción Social"),
        ("Ricardo Rubén Iglesias",     "Secretario de Salud"),
        ("Guillermo Rojas",            "Secretario de Lima"),
    ],
    "Pinamar": [
        ("Francisco Orlando",  "Jefe de Gabinete / Servicios Urbanos"),
        ("Victoria Chust",     "Secretaria de Gobierno"),
        ("Francisco Montes",   "Secretario de Hacienda"),
        ("Sebastián Berardone","Secretario de Seguridad"),
        ("Alejandra Apolonio", "Secretaria de Cultura y Educación"),
        ("Lucas Ventoso",      "Secretario de Turismo y Deportes"),
        ("Eduardo D'Agostino", "Secretario de Salud"),
        ("Sebastián Cufari",   "Secretario de Planeamiento Urbano"),
        ("Mariana Franco",     "Secretaria de Desarrollo Social"),
        ("Martín Rapallino",   "Asesor Letrado"),
    ],
    "Coronel Suárez": [
        ("Mauro Moccero",         "Jefe de Gabinete"),
        ("Marcelo Castorina",     "Director de Gestión Cultural y Ceremonial"),
        ("Victoria Cortalezzi",   "Directora de Prensa"),
        ("Iñaki Etcheberry",      "Director de Turismo y Ambiente"),
        ("Fernando Streitenberger","Director de Comunicaciones"),
        ("Beatriz Larumbe",       "Directora de Defensa al Consumidor"),
        ("Anahí Barreneche",      "Directora de Educación / CREUS"),
    ],
    "Escobar": [
        ("Carlos 'Beto' Ramil",    "Secretario General"),
        ("Pablo Ramos",            "Secretario de Gobierno"),
        ("Edgardo Kutner",         "Secretario de Hacienda"),
        ("Juan Manuel Ordóñez",    "Secretario de Salud"),
        ("Diego Benítez",          "Secretario de Planificación e Infraestructura"),
        ("Verónica Sabena",        "Secretaria de Planificación Territorial y Espacios Públicos"),
        ("Nicolás Gaytan",         "Secretario Legal y Técnica"),
        ("María de los Ángeles Goñi", "Secretaria de Desarrollo Social"),
        ("Graciela Cunial",        "Secretaria Contravencional"),
        ("Pablo Giordano",         "Secretario de Ingresos Públicos"),
        ("Andrés Brandani",        "Secretario de Producción"),
        ("Rocío Fernández",        "Secretaria de Seguridad y Prevención Comunitaria"),
        ("Ariel Lovizio",          "Coordinador General de Relaciones Institucionales"),
    ],
    "Tandil": [
        ("Julio Elichiribehety",   "Jefe de Gabinete"),
        ("Miguel Ángel Lunghi (h)","Vicejefe de Gabinete"),
        ("Alejandra Marcieri",     "Secretaria de Gobierno"),
        ("Luciano Laffose",        "Secretario de Planeamiento y Obras Públicas"),
        ("Juliana Teerink",        "Secretaria de Desarrollo Humano y Hábitat"),
        ("Claudio Biset",          "Secretario de Economía y Administración"),
        ("Marcela Petrantonio",    "Secretaria de Desarrollo Productivo y Relaciones Internacionales"),
        ("Vanesa Frías",           "Secretaria de Protección Ciudadana"),
        ("Javier López",           "Secretario de Legal y Técnica"),
        ("Cecilia Martens",        "Presidenta del Sistema Integrado de Salud"),
    ],
    "Junín": [
        ("Manuel Llovet",       "Secretario General"),
        ("Agustina De Miguel",  "Secretaria de Gobierno"),
        ("Lisandro Benito",     "Secretario de Seguridad"),
        ("Gabriel D'Andrea",    "Secretario de Salud"),
        ("Perla Casella",       "Secretaria de Coordinación"),
        ("Franco Castellazzi",  "Secretario de Obras Públicas y Planificación"),
        ("Eduardo Albarello",   "Secretario de Desarrollo Económico"),
        ("Lorena Linguido",     "Secretaria de Finanzas y Hacienda"),
        ("Melina Fiel",         "Secretaria de Desarrollo Humano"),
    ],
    "Azul": [
        ("Ignacio Pallia",         "Secretario de Gobierno y Jefe de Gabinete"),
        ("Nicolás Tumminaro",      "Secretario de Desarrollo Productivo Local, Asentamiento y Empleo"),
        ("Cecilia Martínez",       "Secretaria de Obras y Servicios Públicos"),
        ("Hernán Combessies",      "Secretario de Salud y Ambiente"),
        ("Laura Barbalarga",       "Secretaria de Educación y Cultura"),
        ("Agustín Carús",          "Secretario de Economía y Finanzas"),
    ],
    "Balcarce": [
        ("Ricardo Heberto Stoppani", "Secretario de Gobierno"),
        ("Francisco Martín Ridao",   "Secretario de Hacienda"),
        ("Luis Gustavo Torres",      "Secretario de Planeamiento, Obras y Servicios Públicos"),
        ("Paola Moreno",             "Secretaria de Desarrollo Social"),
    ],
    "Bolívar": [
        ("Marcos Beorlegui",   "Secretario de Gobierno"),
        ("Javier Erreca",      "Secretario de Finanzas"),
        ("Cecilia Luna",       "Secretaria de Salud"),
        ("Sonia Martínez",     "Secretaria de Desarrollo Social"),
        ("Eduardo Vidarte",    "Secretario de Espacios Públicos"),
        ("Magalí Tullio",      "Secretaria de Relaciones con la Comunidad"),
    ],
    "Bragado": [
        ("Esteban Abel Burga",    "Secretario de Gobierno"),
        ("Emmanuel Aramendi",     "Secretario de Desarrollo Humano y Productivo"),
        ("Alexis Aldo Camús",     "Secretario de Relaciones Institucionales y Caminos Rurales"),
        ("Juan Manuel Santiago",  "Secretario de Obras Públicas"),
        ("Emma Elizalde",         "Secretaria de Salud"),
        ("Nicolás Fernández",     "Secretario de Desarrollo Urbano y Ambiental"),
        ("Juan Manuel Barenghi",  "Secretario de Hacienda"),
        ("José Luis Fernández",   "Subsecretario de Servicios Públicos"),
    ],
    "Cañuelas": [
        ("Mauricio Petre",           "Secretario de Gobierno"),
        ("Fernando Jantus",          "Secretario de Gestión Municipal y Seguridad"),
        ("Simón Gómez",              "Secretario de Desarrollo Social, Niñez y Adolescencia"),
        ("Cintya Curti",             "Secretaria de Educación"),
        ("Hernán Brignani",          "Secretario de Ingresos Públicos"),
        ("Sebastián Demicheli",      "Secretario Legal y Técnico"),
        ("Romina Marques Antunes",   "Secretaria de Presupuesto, Hacienda, Economía y Finanzas"),
        ("Juan Ángel Cruz",          "Secretario de Producción"),
        ("Daniel Arfus",             "Secretario de Salud"),
        ("Gabriela Cabrera",         "Secretaria de Mujeres, Igualdad de Género y Diversidad"),
        ("Alex Goldman",             "Secretario de Relaciones Institucionales y Discapacidad"),
        ("Agustín Lespada",          "Secretario de Medio Ambiente"),
        ("Horacio Endara",           "Secretario de Deportes"),
        ("Marcelo Di Giácomo",       "Secretario de Desarrollo Turístico"),
    ],
    "Chacabuco": [
        ("Javier Gastón Estévez",  "Secretario de Gobierno"),
        ("Natalia Garraza",        "Secretaria de Hacienda"),
        ("Gustavo Masci",          "Secretario General de Salud"),
    ],
    "Chivilcoy": [
        ("Ezequiel Pinotti",            "Secretario de Gobierno"),
        ("Eduardo De Lillo",            "Secretario de Hacienda"),
        ("María del Carmen Ruggirello", "Secretaria de Cultura y Educación"),
        ("Viviana Lemme",               "Secretaria de Desarrollo Social"),
        ("José Luis Neme",              "Secretario de Salud"),
    ],
    "Coronel Dorrego": [
        ("María Laura Dumrauf",  "Secretaria de Gobierno y Hacienda"),
        ("Ramiro Zarzoso",       "Secretario de Ambiente, Obras y Servicios Públicos"),
    ],
    "Coronel Pringles": [
        ("Darío Néstor Christensen", "Jefe de Gabinete"),
        ("Sergio Jackson",           "Secretario de Gobierno y Seguridad"),
        ("Oscar Rossi",              "Secretario de Desarrollo"),
    ],
    "Coronel Rosales": [
        ("Gustavo Rodrigo Sartori",    "Secretario de Gabinete"),
        ("Dr. Jesús César Equiza",     "Secretario de Salud"),
        ("Miriam Riat",                "Secretaria de Obras y Servicios Públicos"),
        ("Mariana Rubio",              "Secretaria de Desarrollo Comunitario"),
        ("José Obregón",               "Secretario de Desarrollo, Producción y Gestión Académica"),
        ("Gabriel Francisco Álvarez",  "Secretario de Seguridad"),
        ("Daniela Sablich",            "Secretaria de Economía"),
    ],
    "Chascomús": [
        ("Cipriano Pérez del Cerro",  "Secretario de Gobierno"),
    ],
    "Dolores": [
        ("Muriel Cifre",     "Secretaria de Salud"),
        ("Virginia Mapelli", "Secretaria de Deportes"),
        ("Agustina Balo",    "Secretaria de Hacienda"),
    ],
    "General Alvarado": [
        ("María Eugenia Bove",     "Secretaria de Gobierno"),
        ("Martín Scarpignato",     "Secretario de Planificación, Ambiente y Obras"),
    ],
    "Lincoln": [
        ("Ramón Parera",   "Jefe de Gabinete"),
    ],
    "Malvinas Argentinas": [
        ("María Luján Salgado",  "Secretaria de Gobierno"),
    ],
    "Marcos Paz": [
        ("Victoria Morel",          "Secretaria de Gobierno y Hábitat"),
        ("Pablo Irrazabal",         "Secretario de Obras Públicas"),
        ("Virginia Goyeneche",      "Secretaria de Economía y Producción"),
        ("Verónica Mc Loughlin",    "Secretaria de Coordinación"),
        ("Viviana Mignani",         "Secretaria de Gestión Pública"),
        ("Silvano Pestrín",         "Secretario de Seguridad"),
        ("Alejandra Tacconi",       "Secretaria de Desarrollo Social"),
        ("Eliana Mabel Martín",     "Secretaria Legal y Técnica"),
        ("Paula McLoughlin",        "Secretaria de Niñez"),
        ("Sabrina Mc Cubbin",       "Secretaria Privada y Ambiente"),
        ("Natalia Gomez",           "Secretaria de Relaciones con la Comunidad"),
        ("Alberto Issouribehere",   "Secretario de Asuntos Municipales y General"),
        ("Adriana Ruíz",            "Secretaria de DDHH y Género"),
        ("Juan Aranda",             "Secretario de Trabajo"),
    ],
    "Mercedes": [
        ("María Clara Zunino",     "Secretaria de Gobierno"),
        ("José Pisano",            "Secretario de Educación"),
        ("David Valerga",          "Secretario de Economía y Hacienda"),
        ("Matías Maresca",         "Secretario de Seguridad"),
        ("Emanuel Pérez Carreras", "Secretario de Obras y Servicios Públicos"),
        ("Néstor Pisapia",         "Secretario de Salud"),
        ("Jorgelina Silva",        "Secretaria de Desarrollo de la Comunidad"),
    ],
    "Monte Hermoso": [
        ("David Quintana",          "Jefe de Gabinete"),
        ("Ramiro Busca",            "Secretario de Gobierno"),
        ("Nerina Dhers",            "Secretaria de Deportes"),
        ("Roberto Pizzato",         "Secretario de Planificación"),
        ("Adrián Ruiz",             "Secretario de Salud"),
        ("Fernando Vera",           "Secretario de Seguridad"),
        ("Ariel Siebenhaar",        "Secretario de Servicios Públicos"),
        ("Marcos Fernández",        "Secretario de Obras Públicas"),
        ("Mariano Prieto",          "Secretario Legal y Técnico"),
        ("Matías De Angelis Rizzo", "Secretario de Hacienda"),
        ("José Abraham",            "Secretario de Tercera Edad"),
        ("Franco Gentili",          "Secretario de Turismo y Cultura"),
        ("Rosana Sotelo",           "Secretaria de Desarrollo Humano y Social"),
        ("Claudio Espinosa",        "Secretario de Comunicación"),
        ("Sandy Puleston",          "Secretaria de Ambiente"),
    ],
    "Navarro": [
        ("Jonathan Castellano",  "Jefe de Gabinete"),
        ("Marcelo Azar",         "Secretario de Producción"),
        ("Hernán Ferrari",       "Secretario de Infraestructura y Obras Públicas"),
    ],
    "Necochea": [
        ("Juan Manuel De La Calle",  "Secretario de Planeamiento, Obras y Servicios Públicos"),
    ],
    "Patagones": [
        ("Héctor Ismael Otero",  "Secretario de Gobierno"),
    ],
    "Pehuajó": [
        ("Rosa Ron",          "Secretaria de Gobierno"),
        ("Florencia Cerquetti","Secretaria de Salud"),
        ("Cristina Román",    "Secretaria de Recursos Humanos"),
        ("Carla Reinoso",     "Secretaria de Finanzas"),
    ],
    "Presidente Perón": [
        ("Mariano Amato",      "Secretario de Gobierno"),
        ("Vanina Ibañez",      "Secretaria de Hacienda y Finanzas"),
        ("Hernán Santana",     "Secretario de Planificación Estratégica y Ambiente"),
        ("Juan Carlos Pérez",  "Secretario de Control de Gestión"),
        ("Alejandro Carbone",  "Secretario de Seguridad"),
        ("Ricardo Aquino",     "Secretario de Servicios Públicos"),
    ],
    "Ramallo": [
        ("Marcela Isarra",   "Secretaria de Gobierno"),
        ("Leandro Torri",    "Secretario de Obras Públicas"),
        ("Pablo Wozniak",    "Secretario de Desarrollo Local"),
    ],
    "Saladillo": [
        ("Viviana Beatriz Rodríguez",  "Secretaria de Gobierno"),
    ],
    "San Nicolás": [
        ("Agustina Gruffat",       "Secretaria de Gobierno"),
        ("Miguel Ángel Battaggia", "Secretario de Hacienda"),
        ("Darío Gritti",           "Secretario de Obras Públicas y Servicios"),
        ("Mirna Bottazzi",         "Secretaria de Salud"),
    ],
    "Tres Arroyos": [
        ("Julio Sebastián Federico", "Jefe de Gabinete"),
        ("Paola Salerno",            "Secretaria de Planeamiento Urbano"),
        ("Sergio Garcimuño",         "Secretario de Hacienda"),
    ],
    "Villa Gesell": [
        ("Pedro Iannuzzi",   "Secretario de Obras Públicas y Servicios"),
        ("Paloma Rodríguez", "Secretaria de Cultura, Educación y Deportes"),
        ("Javier Vicente",   "Secretario de Desarrollo Comunitario"),
    ],
    "Arrecifes": [
        ("María Julia Marincovich", "Secretaria de Gobierno"),
        ("Néstor Cardoso",          "Secretario de Hacienda y Finanzas"),
    ],
    "Benito Juárez": [
        ("Jorge Ismael",   "Jefe de Gabinete"),
        ("Sergio Acosta",  "Secretario de Gobierno"),
    ],
    "Brandsen": [
        ("Gastón Poncetta",  "Jefe de Gabinete"),
    ],
    "Exaltación de la Cruz": [
        ("María Luz Bozzani",        "Secretaria de Gobierno"),
        ("María José Longarella",    "Secretaria de Salud y Calidad de Vida"),
        ("Vanesa Russo",             "Secretaria de Desarrollo de la Comunidad"),
        ("Juan Manuel Boubeta Chehin","Secretario de Seguridad"),
    ],
    "Las Flores": [
        ("Mariángeles Basile",      "Secretaria de Desarrollo Humano"),
        ("Lucas Martín Boggini",    "Secretario de Relaciones Institucionales y Gobierno"),
        ("Duilio Leandro Puente",   "Secretario de Economía y Modernización"),
    ],
    "Lobos": [
        ("Pablo F. Hasper",    "Secretario de Gobierno y Coordinación"),
        ("Gonzalo Menestrina", "Secretario de Hacienda y Producción"),
        ("Julio A. Rustom",    "Secretario de Seguridad"),
        ("Carlos Javier Guzmán","Secretario de Servicios Públicos"),
        ("Antonio Carboni",    "Secretario de Desarrollo Urbano y Ambiente"),
    ],
    "Mar Chiquita": [
        ("Facundo Bocchichio",  "Secretario de Gobierno"),
        ("Ludovico Gordon",     "Secretario de Salud"),
        ("Marcela Lallera",     "Secretaria de Economía"),
        ("Gustavo Rivadavia",   "Secretario de Inspección General"),
        ("Diego Ginestra",      "Secretario de Turismo"),
        ("Sebastián Rioja",     "Secretario de Deportes"),
        ("Marcos Ramundi",      "Secretario de Prensa y Comunicación"),
        ("Selva Lescano",       "Secretaria de Desarrollo Social"),
        ("Sergio García",       "Secretario de Producción y Trabajo"),
        ("Virginia Vasquez",    "Secretaria de Niñez, Adolescencia y Tercera Edad"),
    ],
    "Punta Indio": [
        ("Emiliano Perlini",  "Secretario de Desarrollo Humano"),
    ],
    "Rivadavia": [
        ("Darío Culacciatti",  "Asesor Legal"),
        ("José Carabelli",     "Secretario de Obras Públicas"),
        ("Jorge Gayoso",       "Secretario de Salud"),
        ("Edgardo Ustari",     "Director de Servicios Urbanos"),
    ],
    "Rojas": [
        ("Nicolás Scardino",    "Jefe de Gabinete"),
        ("Lucas Dorazio",       "Secretario de Planificación"),
        ("Esteban Di Camilo",   "Secretario de Seguridad y Letrada"),
        ("Ezequiel Restaine",   "Secretario de Desarrollo Productivo"),
    ],
    "San Andrés de Giles": [
        ("Juan Bautista Castaños", "Secretario de Gobierno"),
        ("Fabián Monti",           "Secretario de Obras y Servicios Públicos"),
    ],
    "San Vicente": [
        ("Mabel Rojas",          "Secretaria General"),
        ("Paula Pereyra",        "Secretaria de Gobierno"),
        ("Roberto Vázquez",      "Secretario de Protección Ciudadana"),
        ("Fernando Vitaller",    "Secretario de Relaciones Institucionales"),
        ("Fernando Rodríguez",   "Director de Comunicaciones y Prensa"),
        ("Daniela Lasalle",      "Secretaria de Gestión Descentralizada"),
        ("Diego Distefano",      "Secretario de Planeamiento y Gestión"),
        ("Gonzalo Mechura",      "Secretario de Salud"),
        ("Claudia Ramos",        "Secretaria de Desarrollo Comunitario"),
        ("Eleonora Vázquez",     "Secretaria de Educación"),
        ("Karina Peñalvel",      "Secretaria de Hacienda"),
        ("Néstor Fernández",     "Secretario de Obras Públicas"),
        ("Sebastián Navarro",    "Secretario de Servicios Públicos"),
        ("Leandro Nimo",         "Secretario de Planificación Territorial y Ambiente"),
        ("Luis Genaro",          "Secretario de Cultura y Deportes"),
    ],
    "Tornquist": [
        ("Claudio Pelizza",         "Secretario de Gobierno y Gestión Pública"),
        ("Bruno Leonel Páez",       "Secretario de Obras y Servicios Públicos"),
        ("Daniela María Aguilar",   "Secretaria de Desarrollo"),
        ("Enrique Alessandrini",    "Secretario de Turismo"),
        ("Noelia Alejandra Baier",  "Secretaria de Salud"),
        ("Victor José Antón",       "Secretario de Ambiente y Desarrollo Productivo"),
    ],
    "Florentino Ameghino": [
        ("Andrés Sciutto",          "Secretario de Gobierno y Hacienda"),
        ("María Ignacia Iturrería", "Secretaria de Desarrollo Social"),
        ("Pablo Varrone",           "Secretario de Salud"),
        ("Bruno Zavatarelli",       "Secretario de Obras y Servicios Públicos"),
        ("Francisco Canosa",        "Secretario de Desarrollo Local"),
    ],
    "Magdalena": [
        ("Andrés García",       "Secretario de Hacienda"),
        ("Julia Mazzoni",       "Subsecretaria de Obras y Servicios Públicos"),
        ("Franco Sanirato",     "Secretario de Salud"),
        ("Valentina Curcio",    "Secretaria de Educación, Desarrollo Humano y Familia"),
        ("Delfina Rocca",       "Directora de Familia, Género y Diversidad"),
        ("Milton Torres",       "Director de Cultura y Prensa"),
        ("Raúl Gómez",          "Director de Producción y Turismo"),
        ("Lucas Sanirato",      "Director de Deportes"),
        ("Nicolás Caloni",      "Director de Control Urbano y Seguridad Ciudadana"),
        ("Daniel García",       "Director de Ingresos Públicos y Tributos"),
        ("Federico Sánchez",    "Director de Servicios Públicos"),
    ],
    "Suipacha": [
        ("Sebastián Maldonado",   "Secretario de Gobierno"),
        ("Ignacio Hernández",     "Secretario de Economía y Hacienda"),
        ("Honorio Rossetti",      "Secretario de Salud"),
        ("Bernardo Martínez Moras","Secretario de Obras Públicas"),
        ("Irene Lubriz",          "Secretaria de Desarrollo Económico y Ambiente"),
    ],
    "Ayacucho": [
        ("Hernán Naveyra",    "Jefe de Gabinete"),
        ("Julián Malvestitti","Secretario de Hacienda"),
        ("Mónica Sarasola",   "Secretaria de Desarrollo Humano"),
    ],
    "Daireaux": [
        ("Roberto Serra",               "Jefe de Gabinete"),
        ("Juan Coito",                  "Secretario de Gobierno"),
        ("Ignacio Mendiondo",           "Secretario de Hacienda"),
        ("Claudia Argentina Bartolomé", "Secretaria de Desarrollo de la Comunidad"),
        ("Walter Martín",               "Secretario de Desarrollo Económico, Educación y Empleo"),
    ],
    "General Arenales": [
        ("Alexis Pires",     "Secretario de Gobierno"),
        ("Darío Perchante",  "Secretario de Hacienda"),
        ("Guillermo Apizzato","Secretario de Salud"),
    ],
    "General Villegas": [
        ("César Aníbal Leiva",     "Secretario de Jefatura de Gabinete"),
        ("Fernando Enrique Galli", "Secretario de Gobierno"),
        ("Gisela Paula Bollini",   "Secretaria de Presupuesto, Hacienda y Finanzas"),
        ("Alejandra Matellan",     "Secretaria de Desarrollo y Planificación Territorial"),
        ("Mónica Spertino",        "Secretaria de Promoción y Producción / Ambiente / Comercio"),
        ("Oscar Alejandro Trojaola","Secretario de Desarrollo Social"),
        ("Silvia Álvarez",         "Secretaria de Deportes y Turismo"),
        ("Luis Alberto Plana",     "Secretario de Seguridad"),
        ("Alejo Arias",            "Secretario de Cultura y Educación"),
    ],
    "Hipólito Yrigoyen": [
        ("Martín Arpigiani",   "Secretario de Gobierno y Hacienda"),
        ("Linda Marina López", "Secretaria de Legales"),
        ("Luciano Martín",     "Secretario de Seguridad"),
        ("Augusto Salvador",   "Secretario de Desarrollo Social"),
    ],
    "San Antonio de Areco": [
        ("Miguel Amadeo",         "Jefe de Gabinete"),
        ("Jimena Berardinelli",   "Secretaria de Hacienda"),
        ("Celina Perez Adamo",    "Secretaria de Desarrollo y Producción"),
        ("Sandra Rezolino",       "Secretaria de Educación y Deporte"),
    ],
    "Carmen de Areco": [
        ("Agustín López",   "Jefe de Gabinete"),
        ("Facundo Ceres",   "Secretario de Gobierno"),
    ],
    "Castelli": [
        ("Sergio Matos",  "Secretario de Hábitat y Vivienda"),
    ],
    "General Guido": [
        ("Natalia Góngora",    "Secretaria de Hacienda"),
        ("Leonardo Arrechea",  "Director de Salud"),
        ("Pía Carini",         "Directora de Obras Públicas"),
        ("Virginia Vergara",   "Directora de Acción Social"),
    ],
    "General Las Heras": [
        ("Marcelo Sartori",       "Jefe de Gabinete"),
        ("Graciana Bacigalupo",   "Secretaria de Gobierno"),
    ],
    "Lezama": [
        ("Diego Mongay",    "Secretario de Políticas Sociales y Desarrollo Territorial"),
        ("Daniel Surdo",    "Secretario de Gobierno y Hacienda"),
        ("Ruben Chirizola", "Secretario de Legal, Técnica y RRHH"),
        ("Mercedes Carrera","Secretaria de Salud"),
        ("Patricio Equiza", "Secretario de Obras y Servicios Públicos"),
    ],
    "Roque Pérez": [
        ("Fabio Britos",          "Secretario General"),
        ("Hernán Reinero",        "Secretario de Planificación, Obras y Servicios"),
        ("Daniel López",          "Secretario Legal y Técnico"),
        ("Ricardo Luis Rinaldi",  "Secretario de Producción, Empleo y Turismo"),
    ],
    "Salliqueló": [
        ("Rosalba Anabela Vargas", "Secretaria de Gobierno"),
        ("Thiago García Blanco",   "Secretario de Hacienda"),
    ],
    "Tapalqué": [
        ("Hebe Bianchi",  "Secretaria de Cultura y Educación"),
    ],
    "Adolfo Alsina": [
        ("Martín Gavio",       "Secretario de Gobierno"),
        ("Renzo Redel",        "Secretario de Hacienda"),
        ("Mariano Viberti",    "Secretario de Obras y Servicios Públicos"),
        ("Facundo Olavarria",  "Secretario de Seguridad"),
        ("Paola Freites",      "Secretaria de Desarrollo Humano"),
        ("Nicolás López",      "Secretario de Turismo"),
        ("Carmen Rantucho",    "Secretaria de Salud"),
        ("Gisele Caussanel",   "Secretaria de Cultura y Educación"),
    ],
    "General Alvear": [
        ("Pedro Gallo",         "Jefe de Gabinete"),
        ("Fabio Gómez Parra",   "Secretario de Gobierno"),
        ("Jorge De Nardi",      "Secretario de Hacienda y Administración"),
        ("Emiliano Barbero",    "Secretario de Obras y Servicios Públicos"),
        ("Hugo Molina",         "Secretario de Desarrollo Económico y Promoción"),
    ],
    "Saavedra": [
        ("Osmar Cleri Herrera", "Secretario de Gobierno y Hacienda"),
        ("Oscar Gomez",         "Secretario de Salud"),
        ("Leticia Iacovelli",   "Secretaria de Cultura y Educación"),
        ("Fernando Couderc",    "Secretario de Desarrollo Económico y Turismo"),
        ("Cecilia Solay",       "Secretaria de Desarrollo Social y Comunitario"),
    ],
    "Adolfo Gonzales Chaves": [
        ("Rodrigo Echayre",              "Secretaría de Gobierno"),
        ("Evelina Pogorzelsky",          "Secretaría de Hacienda"),
        ("Lucía Caprile",                "Secretaría Legal, Técnica y Administrativa"),
        ("Micaela Carolina Vellini",     "Secretaría de Obras y Servicios Públicos"),
        ("Cinthia Fabiana Cancelarich",  "Secretaría de Salud"),
        ("Pablo Arévalo",                "Secretaría de Seguridad"),
    ],
    "General Viamonte": [
        ("Ana Paula Cascallar",   "Secretaría de Gobierno y Hacienda"),
        ("Víctor Amante",         "Secretaría de Relaciones Institucionales y Seguridad Ciudadana"),
        ("Ana María Iribarren",   "Secretaría Legal y Técnica"),
        ("Santiago Catalán Pellet","Secretaría de Salud"),
        ("Carlos Tófoli",         "Secretaría de Obras Públicas"),
        ("Luciana Iribarren",     "Secretaría de Desarrollo Social"),
        ("José Tellería",         "Secretaría de Desarrollo Económico"),
    ],
    "San Cayetano": [
        ("Juan Altamira",     "Secretario General"),
        ("Melisa Eriksen",    "Secretaría de Gobierno"),
        ("Anyelén Benítez",   "Asesora Legal y Técnica"),
    ],
    "Rauch": [
        ("Juan Arbel",           "Jefe de Gabinete"),
        ("Mario Ibargoyen",      "Secretaría de Planificación y Gestión"),
        ("Jimena Dellagogna",    "Secretaría de Obras Públicas"),
        ("Silvina Cabrera",      "Secretaría de Desarrollo Humano y Familiar"),
        ("Valeria Cejas",        "Secretaría de Seguridad y Protección Ciudadana"),
        ("Daniel Poffer",        "Secretaría de Desarrollo Productivo"),
        ("Selva Espondaburu",    "Secretaría de Turismo"),
    ],
    "Pila": [
        ("Guillermo Marchi",             "Jefe de Gabinete"),
        ("María Gabriela Gonzalez Taverna","Secretaría de Gobierno"),
        ("Martín Funes",                 "Secretaría de Hacienda"),
        ("Fabián Amaya",                 "Secretaría de Obras Públicas"),
        ("Karina Pantano",               "Secretaría de Planeamiento y Vivienda"),
        ("Mariana Duarte",               "Secretaría de Salud"),
    ],
    "Villarino": [
        ("Carlos Augusto Morelli",  "Secretaría de Coordinación, Planeamiento y Control de Gestión"),
        ("Claudio Pastori",         "Secretaría de Salud"),
    ],
    "General Madariaga": [
        ("Silvia Mara Simiele",    "Secretaría de Gobierno"),
        ("Emilio José Gomory",     "Secretaría de Hacienda"),
        ("Hernán Ferraro",         "Secretaría Legal, Técnica y Administrativa"),
        ("Javier Alejandro Volpati","Secretaría de Producción"),
        ("Roberto Amadeo Echeverría","Secretaría de Salud"),
        ("Enrique Alberto de Mare", "Secretaría de Desarrollo Social"),
        ("Marcelo Roberto López",  "Secretaría de Seguridad"),
        ("Diana Markovic",         "Secretaría de Obras y Servicios Públicos"),
        ("Karina Uribe",           "Secretaría de Cultura, Turismo y Educación"),
        ("Emiliano San Martín",    "Secretaría de Deportes"),
    ],
    "Carlos Casares": [
        ("Christian Massone",    "Secretaría de Gobierno"),
        ("Marcelo Agradi",       "Secretaría de Hacienda"),
        ("Vanina Gandini",       "Secretaría de Salud y Desarrollo Social"),
        ("Yanile Flavia Mensi",  "Secretaría de Infraestructura y Obras Públicas"),
    ],
    "La Costa": [
        ("Alberto Smith",        "Secretaría General de Gobierno y Administración"),
        ("Gustavo Caruso",       "Secretaría de Producción, Empleo y Trabajo"),
        ("Martín Poustis",       "Secretaría de Desarrollo Humano, Cultura y Comunidad"),
        ("Maria Emilia Martin",  "Secretaría de Educación"),
        ("Amancay López",        "Secretaría de Salud"),
        ("Cristian Escudero",    "Secretaría de Turismo"),
        ("Adrián González",      "Secretaría de Ordenamiento Urbano y Control"),
        ("Walter Natalizia",     "Secretaría de Planeamiento, Infraestructura y Medio Ambiente"),
        ("Elizabeth Becht",      "Secretaría de Hacienda, Economía y Planificación Económica"),
        ("Daniela Gimenez",      "Secretaría de Administración, Legal y Técnica"),
    ],
    "Leandro N. Alem": [
        ("Nancy María Grimaldi", "Secretaría de Gobierno"),
    ],
    "General La Madrid": [
        ("Mario Simón",       "Secretaría de Gobierno"),
        ("Martina Barraza",   "Secretaría de Desarrollo Humano"),
        ("Gastón Álvarez",    "Secretaría de Obras Públicas"),
        ("Mario Conlon",      "Secretaría de Seguridad"),
        ("Domingo Cárceles",  "Secretaría de Salud"),
        ("Carlos Morante",    "Secretaría de Desarrollo Local"),
    ],
    "Pellegrini": [
        ("Bautista Elorza",  "Secretaría de Gobierno"),
    ],
    "Carlos Tejedor": [
        ("Ana Alvez de Mello",      "Secretaría de Gobierno"),
        ("Luciano Bernini",         "Secretaría de Obras Públicas"),
        ("Giuliana Mandrini Gatti", "Secretaría de Servicios Públicos"),
        ("Mariela González",        "Secretaría de Salud"),
        ("Juan Manuel Rojo",        "Secretaría de Seguridad"),
        ("Verónica Porras",         "Secretaría de Desarrollo Humano y Comunidad"),
        ("Paola Pedemonte",         "Secretaría de Hacienda"),
    ],
    "Colón": [
        ("Natalia Y. Avetta",  "Secretaría General de Gobierno y Coordinación"),
    ],
    "Monte": [
        ("Miguel Nasisi",      "Secretaría de Desarrollo Económico Sustentable"),
        ("Matías Balsamello",  "Secretaría de Gestión Institucional"),
    ],
    "Laprida": [
        ("Cristian Sampayo",   "Secretaría de Gobierno"),
        ("Sergio Ferraro",     "Secretaría de Economía y Finanzas"),
        ("Juan Errobidart",    "Secretaría de Infraestructura, Viviendas y Servicios Públicos"),
        ("Fernando Vecini",    "Secretaría de Salud"),
        ("Daniel Bayonés",     "Secretaría de Protección Ciudadana y Orden Público"),
    ],
    "Tres Lomas": [
        ("Gimena Vallejo",          "Secretaría de Gobierno"),
        ("Sebastián Ángel García Rey","Secretaría de Salud y Hospital Municipal"),
    ],
    "Guaminí": [
        ("Gustavo Prienza",     "Jefe de Gabinete"),
        ("Carlos Mariano Fahn", "Secretaría de Obras y Servicios Públicos"),
    ],
    "Veinticinco de Mayo": [
        ("Mercedes Squillaci",           "Secretaría de Gobierno"),
        ("Sergio Mansilla",              "Secretaría de Hacienda"),
        ("Luciano Cabrera",              "Secretaría de Infraestructura y Servicios Públicos"),
        ("María Eugenia Mangialavori",   "Secretaría de Desarrollo Económico y Producción"),
    ],
    "Nueve de Julio": [
        ("Joselina Rodríguez",   "Jefa de Gabinete"),
        ("Víctor Altare",        "Secretaría de Gobierno y Producción"),
        ("María Alejandra Márquez","Secretaría de Desarrollo Comunitario"),
        ("Tamara Vázquez Lagorio","Secretaría de Salud"),
        ("Víctor Bordone",       "Secretaría de Hacienda"),
        ("Juan Pablo Boufflet",  "Secretaría de Obras y Servicios Públicos"),
        ("Martín Banchero",      "Secretaría de Vivienda y Urbanismo"),
    ],
    "Maipú": [
        ("Nahuel Laportilla",    "Secretaría de Gobierno"),
        ("Maximiliano Vecchio",  "Secretaría de Economía y Hacienda"),
    ],
    "Campana": [
        ("Elisa Abella",         "Secretaría de Desarrollo Social y Educación"),
        ("Cecilia Acciardi",     "Secretaría de Salud y Desarrollo Humano"),
        ("Abel Milano",          "Secretaría de Seguridad Ciudadana"),
        ("Fabio Hernández",      "Secretaría de Espacio Público"),
        ("Sergio Agostinelli",   "Secretaría de Planeamiento, Obras y Medio Ambiente"),
        ("Julio Olivastri",      "Secretaría de Hacienda"),
        ("Sergio Roses",         "Secretaría de Desarrollo Económico"),
    ],
    "General Pinto": [
        ("Juan Manuel Blanco",  "Secretario Municipal"),
    ],
    "Alberti": [
        ("Daniel Palazzo",  "Secretaría Legal y Técnica"),
    ],
    "Puán": [
        ("Omar García",  "Secretaría de Desarrollo Social"),
    ],
    "Lobería": [
        ("Augusta Lahore",        "Secretaría de Gobierno"),
        ("M. Magdalena De Noia",  "Secretaría de Economía y Hacienda"),
        ("Daniel Mason",          "Secretaría de Infraestructura, Obras y Servicios Públicos"),
        ("Pedro Barrientos",      "Secretaría de Salud y Desarrollo Social"),
        ("Belén Goyhenespe",      "Secretaría de Innovación y Coordinación"),
    ],
    "General Belgrano": [
        ("Héctor Daniel Arias",  "Secretaría de Gobierno"),
        ("Enrique Acha",         "Secretaría de Hacienda"),
        ("Ramiro Bonini",        "Secretaría de Desarrollo Sustentable"),
    ],
    "Capitán Sarmiento": [
        ("Claudia Nerina Canal",  "Secretaría General de Gobierno y Hacienda"),
    ],
    "San Pedro": [
        ("Alfredo Carrasco",   "Jefe de Gabinete"),
        ("Isabel Carrasco",    "Secretaría de Salud"),
        ("Luciano Arias",      "Secretaría de Desarrollo Económico"),
        ("Mariano Brañas",     "Secretaría de Obras Públicas"),
        ("Adrián Devito",      "Secretaría de Servicios Públicos"),
    ],
    "Baradero": [
        ("Iván Nicolás Moreira",  "Secretaría de Gobierno"),
        ("Darío Puede",           "Secretaría de Salud"),
        ("Antonio Lacera",        "Secretaría de Seguridad"),
        ("Silvana Iozzia",        "Secretaría de Cultura, Turismo, Educación y Deportes"),
    ],
    "Berazategui": [
        ("Antonio Amarilla",       "Secretaría de Gobierno"),
        ("Santiago Castagno",      "Secretaría de Economía"),
        ("Federico López",         "Secretaría de Cultura y Educación"),
        ("María Laura Lacava",     "Secretaría de Desarrollo Social, Recreación, Turismo y Deportes"),
        ("Carlos Balor",           "Secretaría de Obras Públicas"),
        ("Mariel Mussi",           "Secretaría de Salud"),
        ("Sergio Faccenda",        "Secretaría de Servicios Públicos"),
    ],
    "General Lavalle": [
        ("Sebastián Oribe",   "Secretaría de Hacienda y Administración"),
        ("Carlos Acosta",     "Secretaría de Integración Territorial"),
        ("Juan Jaime",        "Secretaría de Desarrollo"),
        ("Danilo Darol",      "Secretaría de Infraestructura"),
    ],
    "General Paz": [
        ("Noelia Massaccesi",      "Secretaría General de Gobierno y Administración"),
        ("María Antonela Duchini", "Secretaría de Economía"),
        ("Marcelo Gastón Aguiar",  "Secretaría de Turismo"),
        ("Adrián Giménez",         "Secretaría de Servicios Públicos"),
        ("Alejandro Iena",         "Secretaría de Cultura"),
        ("Paula Barrera",          "Secretaría de Promoción del Desarrollo y Medio Ambiente"),
        ("Rita Lujan Castro",      "Secretaría de Planificación Social"),
        ("Luciano Lorenzín",       "Secretaría de Obras Públicas y Privadas"),
        ("Alejandra Savino",       "Secretaría de Familia, Niñez y Adolescencia"),
    ],
    "General Rodríguez": [
        ("Luciano Larralde",  "Jefe de Gabinete"),
    ],
    "Salto": [
        ("Camilo Alessandro",  "Secretaría General"),
        ("Alberto Pérez",      "Secretaría de Gobierno"),
        ("Lucas Ale",          "Secretaría de Salud"),
    ],
}

# ─────────────────────────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────────────────────────

COLOR_HEADER      = "1F3864"  # azul oscuro
COLOR_UP          = "E8F5E9"  # verde claro
COLOR_JXC         = "E3F2FD"  # azul claro
COLOR_VECINAL     = "FFF9C4"  # amarillo claro
COLOR_ROW_ALT     = "F5F5F5"  # gris claro
COLOR_SEC_HEADER  = "2E4057"  # azul medio

thin = Side(border_style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def header_style(cell, color=COLOR_HEADER):
    cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

def data_style(cell, bg=None, bold=False, wrap=False):
    cell.font = Font(name="Calibri", bold=bold, size=10)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)
    cell.border = border

def get_color(partido):
    if "Unión por la Patria" in partido:
        return COLOR_UP
    elif "Juntos por el Cambio" in partido:
        return COLOR_JXC
    else:
        return COLOR_VECINAL

# ─────────────────────────────────────────────────────────────
# WORKBOOK
# ─────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()

# ── HOJA 1: Intendentes ──────────────────────────────────────
ws1 = wb.active
ws1.title = "Intendentes"

ws1.row_dimensions[1].height = 30
headers1 = ["#", "Municipio", "Intendente/a", "Partido / Orientación Política", "Secretarios disponibles"]
col_widths1 = [5, 28, 30, 38, 20]

for col, (h, w) in enumerate(zip(headers1, col_widths1), 1):
    cell = ws1.cell(row=1, column=col, value=h)
    header_style(cell)
    ws1.column_dimensions[get_column_letter(col)].width = w

INTENDENTES_sorted = sorted(INTENDENTES, key=lambda x: x[0])

for idx, (municipio, intendente, partido) in enumerate(INTENDENTES_sorted, 1):
    bg = get_color(partido)
    tiene_sec = "Sí" if municipio in SECRETARIOS else "—"
    row_data = [idx, municipio, intendente, partido, tiene_sec]
    for col, val in enumerate(row_data, 1):
        cell = ws1.cell(row=idx+1, column=col, value=val)
        data_style(cell, bg=(bg if idx % 2 == 0 else None))
        if col == 4:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

# Freeze header
ws1.freeze_panes = "A2"

# Leyenda
leyenda_row = len(INTENDENTES_sorted) + 3
ws1.cell(row=leyenda_row,   column=1, value="Leyenda:").font = Font(bold=True)
ws1.cell(row=leyenda_row+1, column=1, value="UP = Unión por la Patria (Peronismo / Kirchnerismo)")
ws1.cell(row=leyenda_row+2, column=1, value="JXC = Juntos por el Cambio (Radicalismo / PRO / coalición opositora)")
ws1.cell(row=leyenda_row+3, column=1, value="Vecinales = partidos locales sin alineación nacional clara")
ws1.cell(row=leyenda_row+4, column=1, value="Datos al: mandato 2023-2027 (elecciones oct. 2023)")

for r in range(leyenda_row, leyenda_row+5):
    ws1.cell(row=r, column=1).font = Font(name="Calibri", italic=True, size=9, color="555555")

# ── HOJA 2: Secretarios ──────────────────────────────────────
ws2 = wb.create_sheet("Secretarios")

ws2.row_dimensions[1].height = 30
headers2 = ["#", "Municipio", "Intendente/a", "Partido", "Nombre Secretario/a", "Cargo / Secretaría"]
col_widths2 = [5, 26, 28, 36, 35, 52]

for col, (h, w) in enumerate(zip(headers2, col_widths2), 1):
    cell = ws2.cell(row=1, column=col, value=h)
    header_style(cell, color=COLOR_SEC_HEADER)
    ws2.column_dimensions[get_column_letter(col)].width = w

# mapa municipio → intendente y partido
mapa = {m: (i, p) for m, i, p in INTENDENTES}

row = 2
counter = 1
for municipio in sorted(SECRETARIOS.keys()):
    intendente, partido = mapa[municipio]
    bg = get_color(partido)
    secretarios_list = SECRETARIOS[municipio]
    for nombre, cargo in secretarios_list:
        row_data = [counter, municipio, intendente, partido, nombre, cargo]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            data_style(cell, bg=(bg if row % 2 == 0 else None), wrap=(col == 6))
        row += 1
        counter += 1

ws2.freeze_panes = "A2"

# Nota al pie
nota_row = row + 2
ws2.cell(row=nota_row, column=1, value=(
    "Nota: Esta hoja contiene los secretarios relevados de sitios oficiales municipales. "
    "Para los 123 municipios restantes se recomienda consultar directamente el sitio web de cada municipio."
)).font = Font(name="Calibri", italic=True, size=9, color="666666")
ws2.merge_cells(f"A{nota_row}:F{nota_row}")
ws2.cell(row=nota_row, column=1).alignment = Alignment(wrap_text=True)
ws2.row_dimensions[nota_row].height = 30

# ── HOJA 3: Resumen ──────────────────────────────────────────
ws3 = wb.create_sheet("Resumen")
ws3.column_dimensions["A"].width = 40
ws3.column_dimensions["B"].width = 15

titulos = [
    ("RESUMEN - INTENDENTES PROVINCIA DE BUENOS AIRES", None),
    ("", None),
    ("Total de municipios", 135),
    ("Unión por la Patria (Peronismo/K)", sum(1 for _,_,p in INTENDENTES if "Unión por la Patria" in p)),
    ("Juntos por el Cambio (Radicalismo/PRO)", sum(1 for _,_,p in INTENDENTES if "Juntos por el Cambio" in p)),
    ("Partidos vecinales / otros", sum(1 for _,_,p in INTENDENTES if "Unión" not in p and "Juntos" not in p)),
    ("", None),
    ("Municipios con secretarios relevados", len(SECRETARIOS)),
    ("Total secretarios cargados", sum(len(v) for v in SECRETARIOS.values())),
    ("", None),
    ("Fuentes", "La Tecla / Infocielo / sitios municipales oficiales"),
    ("Período", "Mandato 2023-2027"),
    ("Fecha de compilación", "Julio 2026"),
]

for r, (label, value) in enumerate(titulos, 1):
    cell_a = ws3.cell(row=r, column=1, value=label)
    if r == 1:
        cell_a.font = Font(name="Calibri", bold=True, size=13, color="1F3864")
    elif value is not None:
        cell_a.font = Font(name="Calibri", size=11)
        cell_b = ws3.cell(row=r, column=2, value=value)
        cell_b.font = Font(name="Calibri", bold=True, size=11)

# ─────────────────────────────────────────────────────────────
# GUARDAR
# ─────────────────────────────────────────────────────────────
output_path = "/home/german/Escritorio/Intendentes_PBA_2023-2027.xlsx"
wb.save(output_path)
print(f"✓ Excel guardado en: {output_path}")
print(f"  - {len(INTENDENTES)} intendentes en hoja 1")
print(f"  - {sum(len(v) for v in SECRETARIOS.values())} secretarios en hoja 2 ({len(SECRETARIOS)} municipios)")
